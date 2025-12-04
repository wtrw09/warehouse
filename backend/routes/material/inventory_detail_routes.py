from fastapi import APIRouter, Depends, HTTPException, Query, Security
from sqlmodel import Session, select, func, and_, or_
from typing import List, Optional
from database import get_db
from core.security import get_current_active_user, get_required_scopes_for_route
from schemas.account.user import UserResponse
from schemas.material.inventory_detail import (
    InventoryDetailResponse,
    PaginatedInventoryDetailsResponse,
    InventoryDetailsListResponse,
    MajorOptionsResponse,
    EquipmentOption,
    EquipmentOptionsResponse
)
from models.material.inventory_detail import InventoryDetail
from models.material.inventory_batch import InventoryBatch
from models.material.material import Material
from models.base.bin import Bin
from models.base.warehouse import Warehouse
from models.base.major import Major
from models.base.equipment import Equipment
from models.base.supplier import Supplier
from schemas.material.batch_code import BatchCodeGenerateRequest, BatchCodeGenerateResponse
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
import io
from datetime import datetime
from fastapi.responses import StreamingResponse
from urllib.parse import quote

router = APIRouter(prefix="/inventory-details", tags=["库存器材明细查询"])


@router.get("", response_model=PaginatedInventoryDetailsResponse, summary="分页查询库存器材明细")
async def get_inventory_details(
    keyword: Optional[str] = Query(None, description="关键词搜索（器材编码、器材名称、规格型号、批次编号、专业名称、装备名称、装备型号）"),
    major_id: Optional[List[int]] = Query(None, description="专业ID数组，支持多选"),
    equipment_id: Optional[List[int]] = Query(None, description="装备ID数组，支持多选"),
    warehouse_id: Optional[int] = Query(None, description="仓库ID"),
    bin_id: Optional[int] = Query(None, description="货位ID"),
    quantity_filter: Optional[str] = Query(None, description="库存数量筛选：'has_stock'（有库存），'no_stock'（无库存），None（全部）"),
    sort_by: str = Query("material_code", description="排序字段"),
    sort_order: str = Query("asc", description="排序方向（asc/desc）"),
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(10, ge=1, le=100, description="每页数量"),
    db: Session = Depends(get_db),
    current_user: UserResponse = Security(get_current_active_user, scopes=get_required_scopes_for_route("/inventory-details"))
):
    """
    分页查询库存器材明细
    
    返回器材ID、名称、规格型号、批次编号、库存数量等信息
    支持通过器材编码、器材查询码、器材名称、规格型号、批次编号、专业名称、装备名称、装备型号进行搜索
    """
    try:
        # 构建基础查询
        query = (
            select(
                InventoryDetail.detail_id,
                InventoryDetail.batch_id,
                Material.id.label("material_id"),
                Material.material_code,
                Material.material_name,
                Material.material_specification.label("material_specification"),
                InventoryBatch.batch_number,
                InventoryDetail.quantity,
                InventoryBatch.unit,
                InventoryBatch.unit_price,
                Supplier.supplier_name,
                InventoryBatch.production_date,
                InventoryBatch.inbound_date,
                Major.id.label("major_id"),
                Major.major_name,
                Equipment.id.label("equipment_id"),
                Equipment.equipment_name,
                Equipment.specification.label("equipment_specification"),
                Bin.bin_name,
                Warehouse.warehouse_name,
                InventoryDetail.last_updated
            )
            .join(InventoryBatch, InventoryDetail.batch_id == InventoryBatch.batch_id)
            .join(Material, InventoryBatch.material_id == Material.id)
            .join(Bin, InventoryDetail.bin_id == Bin.id)
            .join(Warehouse, Bin.warehouse_id == Warehouse.id)
            .join(Major, Material.major_id == Major.id, isouter=True)
            .join(Equipment, Material.equipment_id == Equipment.id, isouter=True)
            .join(Supplier, InventoryBatch.supplier_id == Supplier.id, isouter=True)
        )
        
        # 应用查询条件
        if keyword and keyword.strip():
            keywords = [k.strip().lower() for k in keyword.split() if k.strip()]
            
            if keywords:
                all_keyword_conditions = []
                
                for keyword in keywords:
                    keyword_conditions = []
                    
                    # 搜索器材编码
                    keyword_conditions.append(Material.material_code.ilike(f"%{keyword}%"))
                    
                    # 搜索器材查询码
                    keyword_conditions.append(Material.material_query_code.ilike(f"%{keyword}%"))
                    
                    # 搜索器材名称
                    keyword_conditions.append(Material.material_name.ilike(f"%{keyword}%"))
                    
                    # 搜索器材规格型号
                    keyword_conditions.append(Material.material_specification.ilike(f"%{keyword}%"))
                    
                    # 搜索批次编号
                    keyword_conditions.append(InventoryBatch.batch_number.ilike(f"%{keyword}%"))
                    
                    # 搜索专业名称
                    keyword_conditions.append(Major.major_name.ilike(f"%{keyword}%"))
                    
                    # 搜索装备名称
                    keyword_conditions.append(Equipment.equipment_name.ilike(f"%{keyword}%"))
                    
                    # 搜索装备型号
                    keyword_conditions.append(Equipment.specification.ilike(f"%{keyword}%"))
                    
                    if keyword_conditions:
                        all_keyword_conditions.append(or_(*keyword_conditions))
                
                if all_keyword_conditions:
                    query = query.where(and_(*all_keyword_conditions))
        
        if major_id:
            query = query.where(Material.major_id.in_(major_id))
        
        if equipment_id:
            query = query.where(Material.equipment_id.in_(equipment_id))
        
        if warehouse_id:
            query = query.where(Warehouse.id == warehouse_id)
        
        if bin_id:
            query = query.where(InventoryDetail.bin_id == bin_id)
        
        # 应用库存数量筛选
        if quantity_filter:
            if quantity_filter == "has_stock":
                query = query.where(InventoryDetail.quantity > 0)
            elif quantity_filter == "no_stock":
                query = query.where(InventoryDetail.quantity == 0)
        
        # 获取总数
        count_query = select(func.count()).select_from(query.subquery())
        total = db.exec(count_query).one()
        
        # 应用排序
        sort_field = sort_by
        
        if sort_order.lower() == "desc":
            query = query.order_by(getattr(Material, sort_field).desc(), InventoryBatch.batch_number.asc())
        else:
            query = query.order_by(getattr(Material, sort_field).asc(), InventoryBatch.batch_number.asc())
        
        # 应用分页
        offset = (page - 1) * page_size
        query = query.offset(offset).limit(page_size)
        

        
        # 执行查询
        results = db.exec(query).all()
        
        # 转换为响应模型
        inventory_details = [
            InventoryDetailResponse(
                detail_id=row.detail_id,
                batch_id=row.batch_id,
                material_id=row.material_id,
                material_code=row.material_code,
                material_name=row.material_name,
                material_specification=row.material_specification,
                batch_number=row.batch_number,
                quantity=row.quantity,
                unit=row.unit,
                unit_price=row.unit_price,
                supplier_name=row.supplier_name,
                production_date=row.production_date,
                inbound_date=row.inbound_date,
                major_id=row.major_id,
                major_name=row.major_name,
                equipment_id=row.equipment_id,
                equipment_name=row.equipment_name,
                equipment_specification=row.equipment_specification,
                bin_name=row.bin_name,
                warehouse_name=row.warehouse_name,
                update_time=row.last_updated,
                last_updated=row.last_updated
            )
            for row in results
        ]
        
        return PaginatedInventoryDetailsResponse(
            total=total,
            page=page,
            page_size=page_size,
            data=inventory_details
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"查询库存器材明细失败: {str(e)}")


@router.get("/all", response_model=InventoryDetailsListResponse, summary="获取全部库存器材明细")
async def get_all_inventory_details(
    keyword: Optional[str] = Query(None, description="关键词搜索（器材编码、器材名称、规格型号、批次编号、专业名称、装备名称、装备型号）"),
    major_id: Optional[List[int]] = Query(None, description="专业ID数组，支持多选"),
    equipment_id: Optional[List[int]] = Query(None, description="装备ID数组，支持多选"),
    warehouse_id: Optional[int] = Query(None, description="仓库ID"),
    bin_id: Optional[int] = Query(None, description="货位ID"),
    sort_by: str = Query("material_code", description="排序字段"),
    sort_order: str = Query("asc", description="排序方向（asc/desc）"),
    db: Session = Depends(get_db),
    current_user: UserResponse = Security(get_current_active_user, scopes=get_required_scopes_for_route("/inventory-details/all"))
):
    """
    获取全部库存器材明细（不分页）
    
    返回所有库存器材的详细信息
    支持通过器材编码、器材查询码、器材名称、规格型号、批次编号、专业名称、装备名称、装备型号进行搜索
    """
    try:
        # 构建查询
        query = (
            select(
                InventoryDetail.detail_id,
                InventoryDetail.batch_id,
                Material.id.label("material_id"),
                Material.material_code,
                Material.material_name,
                Material.material_specification.label("material_specification"),
                InventoryBatch.batch_number,
                InventoryDetail.quantity,
                InventoryBatch.unit,
                InventoryBatch.unit_price,
                Supplier.supplier_name,
                InventoryBatch.production_date,
                InventoryBatch.inbound_date,
                Major.id.label("major_id"),
                Major.major_name,
                Equipment.id.label("equipment_id"),
                Equipment.equipment_name,
                Equipment.specification.label("equipment_specification"),
                Bin.bin_name,
                Warehouse.warehouse_name,
                InventoryDetail.last_updated
            )
            .join(InventoryBatch, InventoryDetail.batch_id == InventoryBatch.batch_id)
            .join(Material, InventoryBatch.material_id == Material.id)
            .join(Bin, InventoryDetail.bin_id == Bin.id)
            .join(Warehouse, Bin.warehouse_id == Warehouse.id)
            .join(Major, Material.major_id == Major.id, isouter=True)
            .join(Equipment, Material.equipment_id == Equipment.id, isouter=True)
            .join(Supplier, InventoryBatch.supplier_id == Supplier.id, isouter=True)
        )
        
        # 应用查询条件
        if keyword and keyword.strip():
            keywords = [k.strip().lower() for k in keyword.split() if k.strip()]
            
            if keywords:
                all_keyword_conditions = []
                
                for keyword in keywords:
                    keyword_conditions = []
                    
                    # 搜索器材编码
                    keyword_conditions.append(Material.material_code.ilike(f"%{keyword}%"))
                    
                    # 搜索器材查询码
                    keyword_conditions.append(Material.material_query_code.ilike(f"%{keyword}%"))
                    
                    # 搜索器材名称
                    keyword_conditions.append(Material.material_name.ilike(f"%{keyword}%"))
                    
                    # 搜索器材规格型号
                    keyword_conditions.append(Material.material_specification.ilike(f"%{keyword}%"))
                    
                    # 搜索批次编号
                    keyword_conditions.append(InventoryBatch.batch_number.ilike(f"%{keyword}%"))
                    
                    # 搜索专业名称
                    keyword_conditions.append(Major.major_name.ilike(f"%{keyword}%"))
                    
                    # 搜索装备名称
                    keyword_conditions.append(Equipment.equipment_name.ilike(f"%{keyword}%"))
                    
                    # 搜索装备型号
                    keyword_conditions.append(Equipment.specification.ilike(f"%{keyword}%"))
                    
                    if keyword_conditions:
                        all_keyword_conditions.append(or_(*keyword_conditions))
                
                if all_keyword_conditions:
                    query = query.where(and_(*all_keyword_conditions))
        
        if major_id:
            query = query.where(Material.major_id.in_(major_id))
        
        if equipment_id:
            query = query.where(Material.equipment_id.in_(equipment_id))
        
        if warehouse_id:
            query = query.where(Warehouse.id == warehouse_id)
        
        if bin_id:
            query = query.where(InventoryDetail.bin_id == bin_id)
        
        # 应用排序
        if sort_order.lower() == "desc":
            query = query.order_by(getattr(Material, sort_by).desc(), InventoryBatch.batch_number.asc())
        else:
            query = query.order_by(getattr(Material, sort_by).asc(), InventoryBatch.batch_number.asc())
        
        # 执行查询
        results = db.exec(query).all()
        
        # 转换为响应模型
        inventory_details = [
            InventoryDetailResponse(
                detail_id=row.detail_id,
                batch_id=row.batch_id,
                material_id=row.material_id,
                material_code=row.material_code,
                material_name=row.material_name,
                material_specification=row.material_specification,
                batch_number=row.batch_number,
                quantity=row.quantity,
                unit=row.unit,
                unit_price=row.unit_price,
                supplier_name=row.supplier_name,
                production_date=row.production_date,
                inbound_date=row.inbound_date,
                major_id=row.major_id,
                major_name=row.major_name,
                equipment_id=row.equipment_id,
                equipment_name=row.equipment_name,
                equipment_specification=row.equipment_specification,
                bin_name=row.bin_name,
                warehouse_name=row.warehouse_name,
                update_time=row.last_updated,
                last_updated=row.last_updated
            )
            for row in results
        ]
        
        return InventoryDetailsListResponse(
            total=len(inventory_details),
            data=inventory_details
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取全部库存器材明细失败: {str(e)}")





@router.get("/major-options", response_model=MajorOptionsResponse, summary="获取库存器材所属专业选项集合")
async def get_major_options_from_inventory(
    db: Session = Depends(get_db),
    current_user: UserResponse = Security(get_current_active_user, scopes=get_required_scopes_for_route("/inventory-details/major-options"))
):
    """
    获取库存器材所属专业选项集合（去重）
    
    返回所有库存器材对应的专业选项集合，包含ID和名称
    """
    try:
        # 查询去重后的专业ID和名称
        query = (
            select(Material.major_id, Major.major_name)
            .distinct()
            .join(InventoryBatch, Material.id == InventoryBatch.material_id)
            .join(InventoryDetail, InventoryBatch.batch_id == InventoryDetail.batch_id)
            .join(Major, Material.major_id == Major.id)
            .where(Material.major_id.isnot(None))
            .where(Major.is_delete != True)
        )
        
        results = db.exec(query).all()
        major_options = [{"id": row.major_id, "major_name": row.major_name} for row in results if row.major_id is not None]
        
        return MajorOptionsResponse(
            data=major_options,
            total_count=len(major_options)
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取专业选项集合失败: {str(e)}")


@router.get("/equipment-options", response_model=EquipmentOptionsResponse, summary="获取库存器材所属装备选项集合")
async def get_equipment_options_from_inventory(
    major_ids: List[int] = Query(None, description="专业ID数组，多个专业ID用逗号分隔，为空则获取全部装备选项"),
    db: Session = Depends(get_db),
    current_user: UserResponse = Security(get_current_active_user, scopes=get_required_scopes_for_route("/inventory-details/equipment-options"))
):
    """
    获取库存器材所属装备选项集合（去重）
    
    根据专业ID数组获取对应的装备选项集合，包含ID和display name，如果数组为空则获取全部装备选项
    """
    try:
        # 查询去重后的装备ID
        query = (
            select(Material.equipment_id)
            .distinct()
            .join(InventoryBatch, Material.id == InventoryBatch.material_id)
            .join(InventoryDetail, InventoryBatch.batch_id == InventoryDetail.batch_id)
            .where(Material.equipment_id.isnot(None))
        )
        
        # 如果传入了专业ID数组，则添加专业ID过滤条件
        if major_ids:
            query = query.where(Material.major_id.in_(major_ids))
        
        results = db.exec(query).all()
        equipment_ids = [row for row in results if row is not None]
        
        # 获取装备详细信息
        equipment_options = []
        for equipment_id in equipment_ids:
            equipment = db.exec(select(Equipment).where(Equipment.id == equipment_id)).first()
            if equipment:
                # 构建显示名称
                display_name = f"{equipment.equipment_name}"
                if equipment.specification:
                    display_name += f" {equipment.specification}"
                
                equipment_options.append(EquipmentOption(
                    id=equipment.id,
                    display_name=display_name
                ))
        
        return EquipmentOptionsResponse(
            data=equipment_options,
            total=len(equipment_options)
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取装备选项集合失败: {str(e)}")


@router.get("/statistics", summary="库存器材明细统计")
async def get_inventory_details_statistics(
    db: Session = Depends(get_db),
    current_user: UserResponse = Security(get_current_active_user, scopes=get_required_scopes_for_route("/inventory-details/statistics"))
):
    """
    库存器材明细统计信息
    
    返回库存器材的总数、分类统计等信息
    """
    try:
        # 统计库存器材总数
        total_materials_query = (
            select(func.count(InventoryDetail.detail_id))
            .select_from(InventoryDetail)
        )
        total_materials = db.exec(total_materials_query).one()
        
        # 统计库存器材种类数
        distinct_materials_query = (
            select(func.count(Material.id.distinct()))
            .select_from(InventoryDetail)
            .join(InventoryBatch, InventoryDetail.batch_id == InventoryBatch.batch_id)
            .join(Material, InventoryBatch.material_id == Material.id)
        )
        distinct_materials = db.exec(distinct_materials_query).one()
        
        # 统计库存总数量
        total_quantity_query = (
            select(func.sum(InventoryDetail.quantity))
            .select_from(InventoryDetail)
        )
        total_quantity = db.exec(total_quantity_query).one() or 0
        
        # 统计仓库数量
        distinct_warehouses_query = (
            select(func.count(Warehouse.id.distinct()))
            .select_from(InventoryDetail)
            .join(Bin, InventoryDetail.bin_id == Bin.id)
            .join(Warehouse, Bin.warehouse_id == Warehouse.id)
        )
        distinct_warehouses = db.exec(distinct_warehouses_query).one()
        
        return {
            "total_materials": total_materials,
            "distinct_materials": distinct_materials,
            "total_quantity": total_quantity,
            "distinct_warehouses": distinct_warehouses,
            "update_time": "统计时间: " + str(db.exec(select(func.now())).one())
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取统计信息失败: {str(e)}")

@router.post("/generate-batch-code", response_model=BatchCodeGenerateResponse)
async def generate_batch_code(
    request: BatchCodeGenerateRequest,
    db: Session = Depends(get_db),
    current_user: UserResponse = Security(get_current_active_user, scopes=get_required_scopes_for_route("/inventory-details/generate-batch-code"))
):
    """生成批次编码（器材编码+日期+3位流水），需要IO_edit权限"""
    
    try:
        # 验证器材ID是否存在
        material = db.get(Material, request.material_id)
        if not material:
            raise HTTPException(status_code=404, detail="器材不存在")
        
        # 获取器材编码（10位）
        material_code = material.material_code
        if len(material_code) > 10:
            material_code = material_code[:10]  # 截取前10位
        elif len(material_code) < 10:
            material_code = material_code.ljust(10, '0')  # 不足10位补0
        
        # 格式化日期为YYYYMMDD
        date_str = request.batch_date.strftime("%Y%m%d")
        
        # 构建批次编码前缀：器材编码 + "-" + 日期
        batch_prefix = f"{material_code}-{date_str}"
        
        # 查询系统中所有以该前缀开头的批次编码
        existing_batches = db.exec(
            select(InventoryBatch.batch_number)
            .where(InventoryBatch.batch_number.like(f"{batch_prefix}%"))
        ).all()
        
        # 提取现有的流水号
        used_serial_numbers = []
        for batch_number in existing_batches:
            # 提取流水号部分（最后3位数字）
            serial_part = batch_number.replace(batch_prefix, "")
            if serial_part.isdigit() and len(serial_part) == 3:
                used_serial_numbers.append(int(serial_part))
        
        # 找到最小的未被使用的流水号
        if not used_serial_numbers:
            # 如果没有现有批次，从001开始
            next_serial = 1
        else:
            # 找到最小的未被使用的流水号
            max_serial = max(used_serial_numbers)
            
            # 检查是否有空缺的流水号
            all_possible_serials = set(range(1, max_serial + 2))  # +2确保包含下一个
            used_set = set(used_serial_numbers)
            available_serials = all_possible_serials - used_set
            
            if available_serials:
                next_serial = min(available_serials)
            else:
                next_serial = max_serial + 1
        
        # 确保流水号不超过999
        if next_serial > 999:
            raise HTTPException(status_code=400, detail="当日该器材批次数量已达上限（999），无法生成新的批次编码")
        
        # 格式化流水号为3位数字
        serial_number = str(next_serial).zfill(3)
        
        # 生成完整的批次编码
        batch_code = f"{batch_prefix}{serial_number}"
        
        return BatchCodeGenerateResponse(
            batch_code=batch_code,
            material_code=material.material_code,
            batch_date=request.batch_date,
            sequence=next_serial
        )
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"生成批次编码失败: {str(e)}")

# 导出路由实例
inventory_details_router = router


@router.get("/export-excel", summary="导出库存器材明细到Excel文件")
async def export_inventory_details_to_excel(
    keyword: Optional[str] = Query(None, description="关键词搜索（器材编码、器材名称、规格型号、批次编号、专业名称、装备名称、装备型号）"),
    major_id: Optional[List[int]] = Query(None, description="专业ID数组，支持多选"),
    equipment_id: Optional[List[int]] = Query(None, description="装备ID数组，支持多选"),
    warehouse_id: Optional[int] = Query(None, description="仓库ID"),
    bin_id: Optional[int] = Query(None, description="货位ID"),
    sort_by: str = Query("material_code", description="排序字段"),
    sort_order: str = Query("asc", description="排序方向（asc/desc）"),
    db: Session = Depends(get_db),
    current_user: UserResponse = Security(get_current_active_user, scopes=get_required_scopes_for_route("/inventory-details/export-excel"))
):
    """
    导出库存器材明细到Excel文件
    
    将库存器材明细数据导出为Excel格式文件，支持搜索和筛选条件
    返回.xlsx格式的文件供用户下载
    """
    try:
        # 使用get_all_inventory_details路由的逻辑获取数据
        query = (
            select(
                InventoryDetail.detail_id,
                InventoryDetail.batch_id,
                Material.id.label("material_id"),
                Material.material_code,
                Material.material_name,
                Material.material_specification.label("material_specification"),
                InventoryBatch.batch_number,
                InventoryDetail.quantity,
                InventoryBatch.unit,
                InventoryBatch.unit_price,
                Supplier.supplier_name,
                InventoryBatch.production_date,
                InventoryBatch.inbound_date,
                Major.id.label("major_id"),
                Major.major_name,
                Equipment.id.label("equipment_id"),
                Equipment.equipment_name,
                Equipment.specification.label("equipment_specification"),
                Bin.bin_name,
                Warehouse.warehouse_name,
                InventoryDetail.last_updated
            )
            .join(InventoryBatch, InventoryDetail.batch_id == InventoryBatch.batch_id)
            .join(Material, InventoryBatch.material_id == Material.id)
            .join(Bin, InventoryDetail.bin_id == Bin.id)
            .join(Warehouse, Bin.warehouse_id == Warehouse.id)
            .join(Major, Material.major_id == Major.id, isouter=True)
            .join(Equipment, Material.equipment_id == Equipment.id, isouter=True)
            .join(Supplier, InventoryBatch.supplier_id == Supplier.id, isouter=True)
        )
        
        # 应用查询条件
        if keyword and keyword.strip():
            keywords = [k.strip().lower() for k in keyword.split() if k.strip()]
            
            if keywords:
                all_keyword_conditions = []
                
                for keyword in keywords:
                    keyword_conditions = []
                    
                    # 搜索器材编码
                    keyword_conditions.append(Material.material_code.ilike(f"%{keyword}%"))
                    
                    # 搜索器材查询码
                    keyword_conditions.append(Material.material_query_code.ilike(f"%{keyword}%"))
                    
                    # 搜索器材名称
                    keyword_conditions.append(Material.material_name.ilike(f"%{keyword}%"))
                    
                    # 搜索器材规格型号
                    keyword_conditions.append(Material.material_specification.ilike(f"%{keyword}%"))
                    
                    # 搜索批次编号
                    keyword_conditions.append(InventoryBatch.batch_number.ilike(f"%{keyword}%"))
                    
                    # 搜索专业名称
                    keyword_conditions.append(Major.major_name.ilike(f"%{keyword}%"))
                    
                    # 搜索装备名称
                    keyword_conditions.append(Equipment.equipment_name.ilike(f"%{keyword}%"))
                    
                    # 搜索装备型号
                    keyword_conditions.append(Equipment.specification.ilike(f"%{keyword}%"))
                    
                    if keyword_conditions:
                        all_keyword_conditions.append(or_(*keyword_conditions))
                
                if all_keyword_conditions:
                    query = query.where(and_(*all_keyword_conditions))
        
        if major_id:
            query = query.where(Material.major_id.in_(major_id))
        
        if equipment_id:
            query = query.where(Material.equipment_id.in_(equipment_id))
        
        if warehouse_id:
            query = query.where(Warehouse.id == warehouse_id)
        
        if bin_id:
            query = query.where(InventoryDetail.bin_id == bin_id)
        
        # 应用排序
        if sort_order.lower() == "desc":
            query = query.order_by(getattr(Material, sort_by).desc(), InventoryBatch.batch_number.asc())
        else:
            query = query.order_by(getattr(Material, sort_by).asc(), InventoryBatch.batch_number.asc())
        
        # 执行查询
        results = db.exec(query).all()
        
        # 创建Excel工作簿（使用openpyxl替代xlwt）
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = '库存器材明细'
        
        # 设置表头样式
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                            top=Side(style='thin'), bottom=Side(style='thin'))
        
        # 设置表头
        headers = [
            '明细ID', '批次ID', '器材ID', '器材编码', '器材名称', '器材规格型号',
            '批次编号', '库存数量', '单位', '单价', '供应商名称', '生产日期',
            '入库日期', '专业ID', '专业名称', '装备ID', '装备名称', '装备型号',
            '货位名称', '仓库名称', '最后更新时间'
        ]
        
        for col, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
            # 设置列宽
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
        
        # 设置数据样式
        data_alignment = Alignment(vertical='center', wrap_text=True)
        
        # 填充数据
        for row_idx, row in enumerate(results, start=2):
            worksheet.cell(row=row_idx, column=1, value=row.detail_id).border = thin_border
            worksheet.cell(row=row_idx, column=2, value=row.batch_id).border = thin_border
            worksheet.cell(row=row_idx, column=3, value=row.material_id).border = thin_border
            worksheet.cell(row=row_idx, column=4, value=row.material_code or '').border = thin_border
            worksheet.cell(row=row_idx, column=5, value=row.material_name or '').border = thin_border
            worksheet.cell(row=row_idx, column=6, value=row.material_specification or '').border = thin_border
            worksheet.cell(row=row_idx, column=7, value=row.batch_number or '').border = thin_border
            worksheet.cell(row=row_idx, column=8, value=row.quantity or 0).border = thin_border
            worksheet.cell(row=row_idx, column=9, value=row.unit or '').border = thin_border
            worksheet.cell(row=row_idx, column=10, value=float(row.unit_price or 0)).border = thin_border
            worksheet.cell(row=row_idx, column=11, value=row.supplier_name or '').border = thin_border
            worksheet.cell(row=row_idx, column=12, value=row.production_date.strftime('%Y-%m-%d') if row.production_date else '').border = thin_border
            worksheet.cell(row=row_idx, column=13, value=row.inbound_date.strftime('%Y-%m-%d') if row.inbound_date else '').border = thin_border
            worksheet.cell(row=row_idx, column=14, value=row.major_id or '').border = thin_border
            worksheet.cell(row=row_idx, column=15, value=row.major_name or '').border = thin_border
            worksheet.cell(row=row_idx, column=16, value=row.equipment_id or '').border = thin_border
            worksheet.cell(row=row_idx, column=17, value=row.equipment_name or '').border = thin_border
            worksheet.cell(row=row_idx, column=18, value=row.equipment_specification or '').border = thin_border
            worksheet.cell(row=row_idx, column=19, value=row.bin_name or '').border = thin_border
            worksheet.cell(row=row_idx, column=20, value=row.warehouse_name or '').border = thin_border
            worksheet.cell(row=row_idx, column=21, value=row.last_updated.strftime('%Y-%m-%d %H:%M:%S') if row.last_updated else '').border = thin_border
        
        # 创建内存文件流
        file_stream = io.BytesIO()
        
        # 使用openpyxl直接保存到内存流，避免临时文件
        workbook.save(file_stream)
        file_stream.seek(0)
        
        # 生成文件名（使用URL编码处理中文字符）
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"库存器材明细_{timestamp}.xlsx"
        # 对文件名进行URL编码以支持中文字符
        encoded_filename = quote(filename, safe='')
        
        # 返回文件流响应
        return StreamingResponse(
            file_stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"导出Excel文件失败: {str(e)}")