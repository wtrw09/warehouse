from fastapi import APIRouter, Depends, Security, HTTPException, UploadFile, File, Query
from fastapi.responses import Response
from sqlmodel import Session, select, func
from typing import List, Optional
from io import BytesIO
import os
from datetime import datetime
import logging
import tempfile
from models.base.supplier import Supplier
from schemas.base.supplier import (
    SupplierCreate, SupplierUpdate, SupplierResponse,
    SupplierQueryParams, SupplierPaginationResult,
    BatchSupplierDelete, SupplierStatistics, SupplierBatchImportResult
)
from schemas.account.user import UserResponse
from core.security import get_current_active_user, get_required_scopes_for_route
from database import get_db
from config.import_config import get_import_config
from utils.import_utils import (
    batch_import_transaction, build_entity_data, validate_entity_data,
    batch_insert_entities
)
from utils.error_file_handler import generate_universal_error_file
from utils.template_utils import download_import_template

supplier_router = APIRouter(tags=["供应商管理"], prefix="/suppliers")

# 配置日志记录器
logger = logging.getLogger(__name__)

# 获取供应商列表（分页）
@supplier_router.get("", response_model=SupplierPaginationResult)
def read_suppliers(
    params: SupplierQueryParams = Depends(),
    db: Session = Depends(get_db),
    current_user: UserResponse = Security(get_current_active_user, scopes=get_required_scopes_for_route("/suppliers"))
):
    """获取供应商列表（需要BASE-read权限）"""
    
    # 构建查询条件
    query = select(Supplier).where(Supplier.is_delete != True)
    
    # 处理多关键词搜索
    if params.search and params.search.strip():
        keywords = params.search.strip().split()
        
        # 每个关键词在多个字段中查找，使用AND逻辑
        for keyword in keywords:
            query = query.where(
                Supplier.supplier_name.ilike(f"%{keyword}%") |  # type: ignore
                Supplier.supplier_city.ilike(f"%{keyword}%") |  # type: ignore
                Supplier.supplier_address.ilike(f"%{keyword}%") |  # type: ignore
                Supplier.supplier_contact.ilike(f"%{keyword}%") |  # type: ignore
                Supplier.supplier_manager.ilike(f"%{keyword}%")  # type: ignore
            )
    
    # 按供应商名称筛选
    if params.supplier_name:
        query = query.where(Supplier.supplier_name.ilike(f"%{params.supplier_name}%"))  # type: ignore
    
    # 按所在城市筛选
    if params.supplier_city:
        query = query.where(Supplier.supplier_city.ilike(f"%{params.supplier_city}%"))  # type: ignore
    
    # 按地址筛选
    if params.supplier_address:
        query = query.where(Supplier.supplier_address.ilike(f"%{params.supplier_address}%"))  # type: ignore
    
    # 按联系方式筛选
    if params.supplier_contact:
        query = query.where(Supplier.supplier_contact.ilike(f"%{params.supplier_contact}%"))  # type: ignore
    
    # 按负责人筛选
    if params.supplier_manager:
        query = query.where(Supplier.supplier_manager.ilike(f"%{params.supplier_manager}%"))  # type: ignore
    
    # 排序
    if params.sort_field == "id":
        if params.sort_asc:
            query = query.order_by(Supplier.id)  # type: ignore
        else:
            query = query.order_by(Supplier.id.desc())  # type: ignore
    elif params.sort_field == "supplier_name":
        if params.sort_asc:
            query = query.order_by(Supplier.supplier_name)  # type: ignore
        else:
            query = query.order_by(Supplier.supplier_name.desc())  # type: ignore
    elif params.sort_field == "create_time":
        if params.sort_asc:
            query = query.order_by(Supplier.create_time)  # type: ignore
        else:
            query = query.order_by(Supplier.create_time.desc())  # type: ignore
    elif params.sort_field == "update_time":
        if params.sort_asc:
            query = query.order_by(Supplier.update_time)  # type: ignore
        else:
            query = query.order_by(Supplier.update_time.desc())  # type: ignore
    else:
        # 如果没有有效的排序字段，默认按id升序排序
        query = query.order_by(Supplier.id)  # type: ignore
    
    # 获取总数
    total_count = db.exec(
        select(func.count()).where(Supplier.is_delete != True)
    ).one()
    
    # 分页
    suppliers = db.exec(
        query.offset((params.page - 1) * params.page_size).limit(params.page_size)
    ).all()
    
    total_pages = (total_count + params.page_size - 1) // params.page_size
    
    return {
        "total": total_count,
        "page": params.page,
        "page_size": params.page_size,
        "total_pages": total_pages,
        "data": suppliers
    }

# 获取供应商统计信息
@supplier_router.get("/statistics", response_model=SupplierStatistics)
def get_supplier_statistics(
    db: Session = Depends(get_db),
    current_user: UserResponse = Security(get_current_active_user, scopes=get_required_scopes_for_route("/suppliers/statistics"))
):
    """获取供应商统计信息（需要BASE-read权限）"""
    
    # 获取总供应商数
    total_count = db.exec(
        select(func.count()).where(Supplier.is_delete != True)
    ).one()
    
    return {
        "total_suppliers": total_count
    }

# 下载供应商导入模板
@supplier_router.get("/import-template")
async def download_supplier_import_template(
    current_user: UserResponse = Security(get_current_active_user, scopes=get_required_scopes_for_route("/suppliers/import-template"))
):
    """下载供应商导入模板文件（需要BASE-edit权限）"""
    return await download_import_template('supplier')

# 调试：检查模板配置
@supplier_router.get("/debug/template-config")
async def debug_template_config(
    current_user: UserResponse = Security(get_current_active_user, scopes=get_required_scopes_for_route("/suppliers/import-template"))
):
    """调试接口：查看供应商模板配置"""
    config = get_import_config('supplier')
    if not config:
        return {"error": "不支持的实体类型: supplier"}
    
    return {
        "entity_name": config.entity_name,
        "entity_key": config.entity_key,
        "template_fields": [
            {
                "key": field.key,
                "label": field.label,
                "type": field.type,
                "required": field.required,
                "example": field.example
            } for field in config.template_fields
        ]
    }



# 下载供应商导入错误文件
# 获取单个供应商
@supplier_router.get("/get/{supplier_id}", response_model=SupplierResponse)
def read_supplier(
    supplier_id: int,
    db: Session = Depends(get_db),
    current_user: UserResponse = Security(get_current_active_user, scopes=get_required_scopes_for_route("/suppliers/get"))
):
    """获取单个供应商信息（需要BASE-read权限）"""
    
    supplier = db.get(Supplier, supplier_id)
    if not supplier or supplier.is_delete:
        raise HTTPException(status_code=404, detail="供应商不存在")
    
    return supplier

# 创建供应商
@supplier_router.post("", response_model=SupplierResponse)
def create_supplier(
    supplier: SupplierCreate,
    db: Session = Depends(get_db),
    current_user: UserResponse = Security(get_current_active_user, scopes=get_required_scopes_for_route("/suppliers/new"))
):
    """创建新供应商（需要BASE-edit权限）"""
    
    # 检查供应商名称是否已存在
    existing_supplier = db.exec(
        select(Supplier).where(
            Supplier.supplier_name == supplier.supplier_name,
            Supplier.is_delete != True
        )
    ).first()
    
    if existing_supplier:
        raise HTTPException(status_code=400, detail="供应商名称已存在")
    
    # 创建新供应商
    db_supplier = Supplier(
        supplier_name=supplier.supplier_name,
        supplier_city=supplier.supplier_city,
        supplier_address=supplier.supplier_address,
        supplier_contact=supplier.supplier_contact,
        supplier_manager=supplier.supplier_manager,
        supplier_level=supplier.supplier_level,
        creator=current_user.username  # 使用当前登录用户的用户名作为创建人
    )
    
    db.add(db_supplier)
    db.commit()
    db.refresh(db_supplier)
    
    return db_supplier

# 更新供应商
@supplier_router.put("/update/{supplier_id}", response_model=SupplierResponse)
def update_supplier(
    supplier_id: int,
    supplier_update: SupplierUpdate,
    db: Session = Depends(get_db),
    current_user: UserResponse = Security(get_current_active_user, scopes=get_required_scopes_for_route("/suppliers/update"))
):
    """更新供应商信息（需要BASE-edit权限）"""
    
    db_supplier = db.get(Supplier, supplier_id)
    if not db_supplier or db_supplier.is_delete:
        raise HTTPException(status_code=404, detail="供应商不存在")
    
    # 如果更新供应商名称，检查是否与其他供应商重名
    if supplier_update.supplier_name and supplier_update.supplier_name != db_supplier.supplier_name:
        existing_supplier = db.exec(
            select(Supplier).where(
                Supplier.supplier_name == supplier_update.supplier_name,
                Supplier.is_delete != True,
                Supplier.id != supplier_id
            )
        ).first()
        
        if existing_supplier:
            raise HTTPException(status_code=400, detail="供应商名称已存在")
    
    # 更新字段
    update_data = supplier_update.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        if value is not None:
            setattr(db_supplier, field, value)
    
    db.add(db_supplier)
    db.commit()
    db.refresh(db_supplier)
    
    return db_supplier

# 删除供应商（软删除）
@supplier_router.delete("/delete/{supplier_id}")
def delete_supplier(
    supplier_id: int,
    db: Session = Depends(get_db),
    current_user: UserResponse = Security(get_current_active_user, scopes=get_required_scopes_for_route("/suppliers/delete"))
):
    """删除供应商（软删除，需要BASE-edit权限）"""
    
    db_supplier = db.get(Supplier, supplier_id)
    if not db_supplier or db_supplier.is_delete:
        raise HTTPException(status_code=404, detail="供应商不存在")
    
    # 软删除
    db_supplier.is_delete = True
    
    db.add(db_supplier)
    db.commit()
    
    return {"message": "供应商删除成功"}

# 批量删除供应商
@supplier_router.post("/batch-delete")
def batch_delete_suppliers(
    batch_delete: BatchSupplierDelete,
    db: Session = Depends(get_db),
    current_user: UserResponse = Security(get_current_active_user, scopes=get_required_scopes_for_route("/suppliers/delete"))
):
    """批量删除供应商（需要BASE-edit权限）"""
    
    success_count = 0
    failed_suppliers = []
    
    for supplier_id in batch_delete.supplier_ids:
        try:
            db_supplier = db.get(Supplier, supplier_id)
            if db_supplier and not db_supplier.is_delete:
                db_supplier.is_delete = True
                db.add(db_supplier)
                success_count += 1
        except:
            failed_suppliers.append(supplier_id)
    
    db.commit()
    
    return {
        "success_count": success_count,
        "failed_count": len(failed_suppliers),
        "failed_suppliers": failed_suppliers,
        "message": f"成功删除 {success_count} 个供应商"
    }

# 批量导入供应商（JSON数据）
@supplier_router.post("/batch-import-data", response_model=SupplierBatchImportResult)
async def batch_import_suppliers_data(
    suppliers_data: List[SupplierCreate],
    db: Session = Depends(get_db),
    current_user: UserResponse = Security(get_current_active_user, scopes=get_required_scopes_for_route("/suppliers/new"))
):
    """批量导入供应商数据（JSON格式，需要BASE-edit权限）"""
    
    try:
        if not suppliers_data:
            raise HTTPException(status_code=400, detail="请提供要导入的供应商数据")
        
        # 获取供应商导入配置
        config = get_import_config('supplier')
        if not config:
            raise HTTPException(status_code=400, detail="不支持的实体类型: supplier")
        
        # 转换数据格式以匹配验证函数期望的格式
        formatted_data = []
        for i, supplier in enumerate(suppliers_data, start=2):  # 从第2行开始（第1行为标题）
            entity_data = {
                'supplier_name': supplier.supplier_name or '',
                'supplier_city': supplier.supplier_city or '',
                'supplier_address': supplier.supplier_address or '',
                'supplier_manager': supplier.supplier_manager or '',
                'supplier_contact': supplier.supplier_contact or '',
                'supplier_level': supplier.supplier_level
            }
            formatted_data.append(entity_data)
        
        # 验证数据
        all_errors = await validate_entity_data(formatted_data, config, db)
        
        # 筛选出合格数据
        valid_data = []
        error_data = []
        
        for i, supplier_data in enumerate(formatted_data):
            row_index = i + 2  # Excel行号
            has_error = any(error.row_index == row_index for error in all_errors)
            if not has_error:
                valid_data.append((row_index, supplier_data))
            else:
                error_data.append((row_index, supplier_data))
        
        # 批量插入合格数据
        success_count = 0
        insert_errors = []
        
        # 如果有错误数据，生成错误Excel文件
        error_file_path = None
        error_file_name = None
        if all_errors:
            from utils.error_file_handler import process_import_errors_and_generate_file
            
            # 使用通用函数处理错误文件生成
            # 构造original_rows（JSON导入的情况下需要模拟）
            original_rows = []
            for i, supplier_data in enumerate(formatted_data):
                row_index = i + 2  # Excel行号
                # 按配置的字段顺序构造行数据
                row_data = tuple(str(supplier_data.get(field.key, '')) for field in config.template_fields)
                original_rows.append((row_index, row_data))
            
            error_file_path, error_file_name = process_import_errors_and_generate_file(
                all_errors=all_errors,
                original_rows=original_rows,
                config=config,
                entity_key='supplier',
                username=current_user.username
            )
        
        with batch_import_transaction(db):
            for row_index, entity_data in valid_data:
                try:
                    # 插入前再次检查唯一性（防止并发导入时的重复）
                    existing_supplier = db.exec(
                        select(Supplier).where(
                            Supplier.supplier_name == entity_data['supplier_name'].strip(),
                            Supplier.is_delete != True
                        )
                    ).first()
                    
                    if existing_supplier:
                        from schemas.common.import_schemas import ImportError
                        insert_errors.append(ImportError(
                            row_index=row_index,
                            field='supplier_name',
                            error_message=f'供应商:供应商名称:"{entity_data["supplier_name"]}"在导入过程中与现有数据冲突',
                            raw_data=entity_data
                        ))
                        continue
                    
                    # 创建供应商记录
                    supplier_data = {
                        'supplier_name': entity_data.get('supplier_name', '').strip(),
                        'supplier_city': entity_data.get('supplier_city', '').strip() or None,
                        'supplier_address': entity_data.get('supplier_address', '').strip() or None,
                        'supplier_manager': entity_data.get('supplier_manager', '').strip() or None,
                        'supplier_contact': entity_data.get('supplier_contact', '').strip() or None,
                        'supplier_level': entity_data.get('supplier_level') or None,
                        'creator': current_user.username
                    }
                    
                    db_supplier = Supplier(**supplier_data)
                    db.add(db_supplier)
                    success_count += 1
                    
                except Exception as e:
                    from schemas.common.import_schemas import ImportError
                    insert_errors.append(ImportError(
                        row_index=row_index,
                        field='database',
                        error_message=f"供应商:数据库:插入数据库时出错: {str(e)}",
                        raw_data=entity_data
                    ))
        
        # 合并所有错误
        all_errors.extend(insert_errors)
        
        # 构建返回结果
        result = SupplierBatchImportResult(
            total_count=len(suppliers_data),
            success_count=success_count,
            error_count=len(all_errors),
            errors=all_errors,
            import_time=datetime.now(),
            has_error_file=error_file_path is not None,
            error_file_name=error_file_name
        )
        
        return result
        
    except HTTPException:
        raise
    except Exception as e:
        # 记录详细错误信息
        import traceback
        error_detail = f"批量导入失败: {str(e)}"
        print(f"Import error: {error_detail}")
        print(f"Traceback: {traceback.format_exc()}")
        
        # 返回错误结果
        from schemas.common.import_schemas import ImportError
        return SupplierBatchImportResult(
            total_count=0,
            success_count=0,
            error_count=1,
            errors=[ImportError(
                row_index=1,
                field="system",
                error_message=error_detail,
                raw_data={}
            )],
            import_time=datetime.now(),
            has_error_file=False
        )

# 批量导入供应商（文件上传）
@supplier_router.post("/batch-import", response_model=SupplierBatchImportResult)
async def batch_import_suppliers(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: UserResponse = Security(get_current_active_user, scopes=get_required_scopes_for_route("/suppliers/new"))
):
    """批量导入供应商数据，不合格数据导出到新文件（需要BASE-edit权限）"""
    print(f"开始处理文件: {file.filename}");
    # 验证文件类型
    if not file.filename:
        raise HTTPException(status_code=400, detail="供应商:文件:请选择要上传的文件")
    
    # 检查文件扩展名
    allowed_extensions = ['.xlsx', '.xls']
    file_extension = None
    for ext in allowed_extensions:
        if file.filename.lower().endswith(ext):
            file_extension = ext
            break
    
    if not file_extension:
        raise HTTPException(
            status_code=400, 
            detail=f"供应商:文件:不支持的文件格式。请上传Excel文件（{', '.join(allowed_extensions)}）"
        )
    
    # 检查MIME类型
    allowed_mime_types = [
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',  # .xlsx
        'application/vnd.ms-excel',  # .xls
        'application/octet-stream'  # 通用二进制格式
    ]
    
    if file.content_type and file.content_type not in allowed_mime_types:
        print(f"警告：上传文件的MIME类型为 {file.content_type}，但将继续尝试处理")
    
    # 获取供应商导入配置
    config = get_import_config('supplier')
    if not config:
        raise HTTPException(status_code=400, detail="供应商:实体类型:不支持的实体类型: supplier")
    
    try:
        # 读取Excel文件
        contents = await file.read()
        
        # 检查文件是否为空
        if not contents:
            raise HTTPException(status_code=400, detail="供应商:文件:上传的文件为空")
        
        # 检查文件大小（最大10MB）
        if len(contents) > 10 * 1024 * 1024:
            raise HTTPException(status_code=400, detail="供应商:文件:文件过大，请确保文件小于10MB")
        
        # 读取Excel文件
        suppliers_data = []
        original_rows = []  # 保存原始行数据
        
        try:
            if file_extension == '.xlsx':
                # 使用openpyxl读取.xlsx文件
                from openpyxl import load_workbook
                workbook = load_workbook(BytesIO(contents))
                sheet = workbook.active
                
                for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):  # type: ignore
                    if row and row[0]:  # 第一个字段不为空
                        entity_data = build_entity_data(row, config)
                        suppliers_data.append((row_index, entity_data))
                        original_rows.append((row_index, row))
                
                print(f"使用openpyxl成功读取.xlsx文件，共{len(suppliers_data)}行数据")
                
            elif file_extension == '.xls':
                # 使用xlrd读取.xls文件
                import xlrd
                workbook = xlrd.open_workbook(file_contents=contents)
                sheet = workbook.sheet_by_index(0)
                
                for row_index in range(1, sheet.nrows):  # 从第2行开始（跳过标题行）
                    row = sheet.row_values(row_index)
                    if row and row[0]:  # 第一个字段不为空
                        entity_data = build_entity_data(row, config)
                        suppliers_data.append((row_index + 1, entity_data))  # Excel行号从1开始
                        original_rows.append((row_index + 1, row))
                
                print(f"使用xlrd成功读取.xls文件，共{len(suppliers_data)}行数据")
            else:
                raise HTTPException(status_code=400, detail=f"供应商:文件:不支持的文件格式: {file_extension}")
            
        except Exception as e:
            print(f"Excel文件读取失败: {str(e)}")
            raise HTTPException(status_code=400, detail=f"供应商:文件:无法读取Excel文件: {str(e)}")
        
        # 如果没有读取到任何数据
        if not suppliers_data:
            raise HTTPException(
                status_code=400, 
                detail="供应商:文件:文件中没有找到有效的数据行。请检查：\n1. 文件是否包含数据（除了标题行）\n2. 第一列（供应商名称）是否填写"
            )

        # 验证数据（包含两步重复性检查）
        all_errors = await validate_entity_data(
            [data for _, data in suppliers_data], 
            config, 
            db
        )
        
        # 筛选出合格数据
        valid_data = []
        error_rows = []
        
        for row_index, entity_data in suppliers_data:
            has_error = any(error.row_index == row_index for error in all_errors)
            if not has_error:
                valid_data.append((row_index, entity_data))
            else:
                # 收集该行的所有错误
                row_errors = [error for error in all_errors if error.row_index == row_index]
                for orig_row_index, orig_row in original_rows:
                    if orig_row_index == row_index:
                        error_rows.append((row_index, orig_row, row_errors))
                        break
        
        # 如果有验证错误，生成错误文件（临时变量，不直接返回）
        temp_error_file_path = None
        if error_rows:
            temp_error_file_path = generate_universal_error_file(error_rows, all_errors, config, entity_key='supplier')
        
        # 批量插入合格数据
        success_count = 0
        insert_errors = []
        
        with batch_import_transaction(db):
            for row_index, entity_data in valid_data:
                try:
                    # 插入前再次检查唯一性（防止并发导入时的重复）
                    existing_supplier = db.exec(
                        select(Supplier).where(
                            Supplier.supplier_name == entity_data['supplier_name'].strip(),
                            Supplier.is_delete != True
                        )
                    ).first()
                    
                    if existing_supplier:
                        from schemas.common.import_schemas import ImportError
                        insert_errors.append(ImportError(
                            row_index=row_index,
                            field='supplier_name',
                            error_message=f'供应商:文件:文件"{entity_data["supplier_name"]}"在导入过程中与现有数据冲突',
                            raw_data=entity_data
                        ))
                        continue
                    
                    # 创建供应商记录
                    # 确保所有必要字段都有值
                    supplier_data = {
                        'supplier_name': entity_data.get('supplier_name', '').strip(),
                        'supplier_city': entity_data.get('supplier_city', '').strip() or None,
                        'supplier_address': entity_data.get('supplier_address', '').strip() or None,
                        'supplier_manager': entity_data.get('supplier_manager', '').strip() or None,
                        'supplier_contact': entity_data.get('supplier_contact', '').strip() or None,
                        'supplier_level': entity_data.get('supplier_level') or None,
                        'creator': current_user.username
                    }
                    
                    db_supplier = Supplier(**supplier_data)
                    db.add(db_supplier)
                    success_count += 1
                    
                except Exception as e:
                    from schemas.common.import_schemas import ImportError
                    insert_errors.append(ImportError(
                        row_index=row_index,
                        field='database',
                        error_message=f"供应商:数据库:插入数据库时出错: {str(e)}",
                        raw_data=entity_data
                    ))
        
        # 如果有插入错误，更新错误文件
        if insert_errors:
            from utils.error_file_handler import add_insert_errors_to_error_list
            # 将ImportError对象转换为字典列表
            insert_errors_dict = [
                {
                    'row_index': error.row_index,
                    'field': error.field,
                    'error_message': error.error_message,
                    'raw_data': error.raw_data
                }
                for error in insert_errors
            ]
            add_insert_errors_to_error_list(insert_errors_dict, all_errors, original_rows, error_rows)
        
        # 生成最终错误文件（包含所有错误：验证错误+插入错误）
        error_file_path = None
        error_file_name = None
        if all_errors:
            from utils.error_file_handler import generate_error_file_from_all_errors
            error_file_path = generate_error_file_from_all_errors(all_errors, original_rows, config, entity_key='supplier')
            if error_file_path:
                # 只返回文件名，前端负责构造下载URL
                error_file_name = os.path.basename(error_file_path)
        #输出错误文件名
        print(f"生成供应商错误文件: {error_file_name}");
        # 构建返回结果
        result = SupplierBatchImportResult(
            total_count=len(suppliers_data),
            success_count=success_count,
            error_count=len(all_errors),
            errors=all_errors,
            import_time=datetime.now(),
            has_error_file=error_file_path is not None,
            error_file_name=error_file_name
        )
        
        return result
        
    except Exception as e:
        # 记录详细错误信息
        import traceback
        error_detail = f"批量导入失败: {str(e)}"
        print(f"Import error: {error_detail}")
        print(f"Traceback: {traceback.format_exc()}")
        
        # 返回错误结果，包含具体错误信息
        from schemas.common.import_schemas import ImportError
        return SupplierBatchImportResult(
            total_count=0,
            success_count=0,
            error_count=1,
            errors=[ImportError(
                row_index=1,
                field="file",
                error_message=error_detail,
                raw_data={}
            )],
            import_time=datetime.now(),
            has_error_file=False
        )

# 下载供应商导入错误文件
@supplier_router.get("/download-error-file")
async def download_supplier_error_file(
    file_name: str = Query(..., description="错误文件名"),
    current_user: UserResponse = Security(get_current_active_user, scopes=get_required_scopes_for_route("/suppliers/download-error-file"))
):
    """下载供应商导入错误数据文件（需要BASE-edit权限）"""
    # 构建默认供应商错误文件目录
    temp_dir = tempfile.gettempdir()
    supplier_dir = os.path.join(temp_dir, 'supplier')
    
    # 直接使用文件名（前端已经进行了URL编码，后端直接使用）
    decoded_file_name = file_name
    
    # 构建完整文件路径
    file_path = os.path.join(supplier_dir, decoded_file_name)
    
    if not os.path.exists(file_path):
        logger.error(f"错误文件不存在: {file_path}")
        raise HTTPException(status_code=404, detail="错误文件不存在或已过期")
    
    try:
        # 读取文件内容 - 使用二进制模式避免编码问题
        with open(file_path, 'rb') as file:
            file_content = file.read()
        
        # 获取文件大小
        file_size = len(file_content)
        
        logger.info(f"文件读取成功: 文件大小={file_size} bytes")
            
        # 检查文件内容是否为空
        if file_size == 0:
            logger.warning(f"文件内容为空: {file_path}")
            raise HTTPException(status_code=404, detail="错误文件为空或损坏")
        
        logger.info("开始返回文件内容")
        
        # 直接返回响应内容
        return Response(
            content=file_content,
            media_type="application/vnd.ms-excel",
            headers={
                "Content-Disposition": f"attachment; filename=\"{decoded_file_name}\"",
                "Content-Length": str(file_size)
            }
        )
        
    except Exception as e:
        logger.error(f"读取文件失败: {file_path}, 错误: {str(e)}", exc_info=True)
        # 提供更详细的错误信息
        error_detail = f"读取文件失败: {str(e)}"
        if "latin-1" in str(e):
            error_detail += "。请检查文件路径是否包含特殊字符。"
        raise HTTPException(status_code=500, detail=error_detail)