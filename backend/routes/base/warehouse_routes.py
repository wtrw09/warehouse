from fastapi import APIRouter, Depends, Security, HTTPException, UploadFile, File, BackgroundTasks, Query
from fastapi.responses import Response
from sqlmodel import Session, select, func
from typing import List, Any
import os
import tempfile
import logging
from io import BytesIO

logger = logging.getLogger(__name__)
from datetime import datetime

from models.base.warehouse import Warehouse
from schemas.base.warehouse import (
    WarehouseCreate, WarehouseUpdate, WarehouseResponse,
    WarehouseQueryParams, WarehousePaginationResult,
    BatchWarehouseDelete, WarehouseStatistics
)
from schemas.base.warehouse import WarehouseBatchImportResult
from schemas.account.user import UserResponse
from schemas.common.import_schemas import ImportError
from core.security import get_current_active_user, get_required_scopes_for_route
from database import get_db
from utils.import_utils import (
    validate_entity_data, batch_insert_entities, batch_import_transaction,
    build_entity_data, get_existing_values
)

from config.import_config import get_import_config
from utils.template_utils import download_import_template

warehouse_router = APIRouter(tags=["仓库管理"], prefix="/warehouses")

# 仓库数据批量导入 - 文件上传方式
@warehouse_router.post("/batch-import", response_model=WarehouseBatchImportResult)
async def batch_import_warehouses(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: UserResponse = Security(get_current_active_user, scopes=get_required_scopes_for_route("/warehouses"))
):
    """批量导入仓库数据，不合格数据导出到新文件（需要BASE-edit权限）"""
    print(f"开始处理文件: {file.filename}")
    # 验证文件类型
    if not file.filename:
        raise HTTPException(status_code=400, detail="仓库:文件:请选择要上传的文件")
    
    # 检查文件扩展名
    allowed_extensions = ['.xlsx', '.xls']
    file_extension = None
    for ext in allowed_extensions:
        if file.filename.lower().endswith(ext):
            file_extension = ext
            break
    
    if not file_extension:
        raise HTTPException(
            status_code=400, 
            detail=f"仓库:文件:不支持的文件格式。请上传Excel文件（{', '.join(allowed_extensions)}）"
        )
    
    # 检查MIME类型
    allowed_mime_types = [
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',  # .xlsx
        'application/vnd.ms-excel',  # .xls
        'application/octet-stream'  # 通用二进制格式
    ]
    
    if file.content_type and file.content_type not in allowed_mime_types:
        print(f"警告：上传文件的MIME类型为 {file.content_type}，但将继续尝试处理")
    
    # 获取仓库导入配置
    config = get_import_config('warehouse')
    if not config:
        raise HTTPException(status_code=400, detail="仓库:配置:不支持的实体类型: warehouse")
    
    try:
        # 读取Excel文件
        contents = await file.read()
        
        # 检查文件是否为空
        if not contents:
            raise HTTPException(status_code=400, detail="仓库:文件:上传的文件为空")
        
        # 检查文件大小（最大10MB）
        if len(contents) > 10 * 1024 * 1024:
            raise HTTPException(status_code=400, detail="仓库:文件:文件过大，请确保文件小于10MB")
        
        # 读取Excel文件
        warehouses_data = []
        original_rows = []  # 保存原始行数据
        
        try:
            if file_extension == '.xlsx':
                # 使用openpyxl读取.xlsx文件
                from openpyxl import load_workbook
                workbook = load_workbook(BytesIO(contents))
                sheet = workbook.active
                
                for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):  # type: ignore
                    if row and row[0]:  # 第一个字段不为空
                        entity_data = build_entity_data(row, config)
                        warehouses_data.append((row_index, entity_data))
                        original_rows.append((row_index, row))
                
                print(f"使用openpyxl成功读取.xlsx文件，共{len(warehouses_data)}行数据")
                
            elif file_extension == '.xls':
                # 使用xlrd读取.xls文件
                import xlrd
                workbook = xlrd.open_workbook(file_contents=contents)
                sheet = workbook.sheet_by_index(0)
                
                for row_index in range(1, sheet.nrows):  # 从第2行开始（跳过标题行）
                    row = sheet.row_values(row_index)
                    if row and row[0]:  # 第一个字段不为空
                        entity_data = build_entity_data(row, config)
                        warehouses_data.append((row_index + 1, entity_data))  # Excel行号从1开始
                        original_rows.append((row_index + 1, row))
                
                print(f"使用xlrd成功读取.xls文件，共{len(warehouses_data)}行数据")
            else:
                raise HTTPException(status_code=400, detail=f"仓库:文件:不支持的文件格式: {file_extension}")
            
        except Exception as e:
            print(f"Excel文件读取失败: {str(e)}")
            raise HTTPException(status_code=400, detail=f"仓库:文件:无法读取Excel文件: {str(e)}")
        
        # 如果没有读取到任何数据
        if not warehouses_data:
            raise HTTPException(
                status_code=400, 
                detail="仓库:文件:文件中没有找到有效的数据行。请检查：\n1. 文件是否包含数据（除了标题行）\n2. 第一列（仓库名）是否填写"
            )

        # 验证数据（包含两步重复性检查）
        all_errors = await validate_entity_data(
            [data for _, data in warehouses_data], 
            config, 
            db
        )
        
        # 筛选出合格数据
        valid_data = []
        error_rows = []
        
        for row_index, entity_data in warehouses_data:
            has_error = any(error.row_index == row_index for error in all_errors)
            if not has_error:
                valid_data.append((row_index, entity_data))
            else:
                # 收集该行的所有错误
                row_errors = [error for error in all_errors if error.row_index == row_index]
                for orig_row_index, orig_row in original_rows:
                    if orig_row_index == row_index:
                        error_rows.append((row_index, orig_row, row_errors))
                        break
        
        # 如果有验证错误，生成错误文件（临时变量，不直接返回）
        temp_error_file_path = None
        if error_rows:
            from utils.error_file_handler import generate_universal_error_file
            print(f"错误行数据: {error_rows}");
            temp_error_file_path = generate_universal_error_file(error_rows, all_errors, config, entity_key='warehouse')
        
        # 批量插入合格数据
        success_count = 0
        insert_errors = []
        
        with batch_import_transaction(db):
            for row_index, entity_data in valid_data:
                try:
                    # 插入前再次检查唯一性（防止并发导入时的重复）
                    existing_warehouse = db.exec(
                        select(Warehouse).where(
                            Warehouse.warehouse_name == entity_data['warehouse_name'].strip(),
                            Warehouse.is_delete != True
                        )
                    ).first()
                    
                    if existing_warehouse:
                        from schemas.common.import_schemas import ImportError
                        insert_errors.append(ImportError(
                            row_index=row_index,
                            field='warehouse_name',
                            error_message=f'仓库:仓库名:"{entity_data["warehouse_name"]}"在导入过程中与现有数据冲突',
                            raw_data=entity_data
                        ))
                        continue
                    
                    # 创建仓库记录
                    # 确保所有必要字段都有值
                    warehouse_data = {
                        'warehouse_name': entity_data.get('warehouse_name', '').strip(),
                        'warehouse_city': entity_data.get('warehouse_city', '').strip() or None,
                        'warehouse_address': entity_data.get('warehouse_address', '').strip() or None,
                        'warehouse_manager': entity_data.get('warehouse_manager', '').strip() or None,
                        'warehouse_contact': entity_data.get('warehouse_contact', '').strip() or None,
                        'warehouse_level': entity_data.get('warehouse_level') or None,
                        'creator': current_user.username
                    }
                    
                    db_warehouse = Warehouse(**warehouse_data)
                    db.add(db_warehouse)
                    success_count += 1
                    
                except Exception as e:
                    from schemas.common.import_schemas import ImportError
                    insert_errors.append(ImportError(
                        row_index=row_index,
                        field='database',
                        error_message=f"仓库:数据库:插入数据库时出错: {str(e)}",
                        raw_data=entity_data
                    ))
        
        # 如果有插入错误，更新错误文件
        if insert_errors:
            from utils.error_file_handler import add_insert_errors_to_error_list
            # 将ImportError对象转换为字典
            insert_errors_dict = [
                {
                    'row_index': error.row_index,
                    'field': error.field,
                    'error_message': error.error_message,
                    'raw_data': error.raw_data
                }
                for error in insert_errors
            ]
            add_insert_errors_to_error_list(insert_errors_dict, all_errors, original_rows, error_rows)
        
        # 生成最终错误文件（包含所有错误：验证错误+插入错误）
        error_file_path = None
        error_file_name = None
        if all_errors:
            from utils.error_file_handler import generate_error_file_from_all_errors
            error_file_path = generate_error_file_from_all_errors(all_errors, original_rows, config, entity_key='warehouse')
            if error_file_path:
                # 只返回文件名，前端需要下载的时候提供文件名
                error_file_name = os.path.basename(error_file_path)
        #输出错误文件名
        print(f"生成仓库错误文件: {error_file_name}")
        
        # 构建返回结果
        return WarehouseBatchImportResult(
            total_count=len(warehouses_data),
            success_count=success_count,
            error_count=len(all_errors),
            errors=all_errors,
            import_time=datetime.now(),
            has_error_file=error_file_path is not None,
            error_file_name=error_file_name
        )
        
    except Exception as e:
        # 记录详细错误信息
        import traceback
        error_detail = f"批量导入失败: {str(e)}"
        print(f"Import error: {error_detail}")
        print(f"Traceback: {traceback.format_exc()}")
        
        # 返回错误结果，包含具体错误信息
        from schemas.common.import_schemas import ImportError
        return WarehouseBatchImportResult(
            total_count=0,
            success_count=0,
            error_count=1,
            errors=[ImportError(
                row_index=1,
                field="file",
                error_message=error_detail,
                raw_data={}
            ).model_dump()],
            import_time=datetime.now(),
            has_error_file=False
        )

# 仓库数据批量导入 - JSON数据方式
@warehouse_router.post("/batch-import-data", response_model=WarehouseBatchImportResult)
async def batch_import_warehouse_data(
    warehouses_data: List[dict],
    db: Session = Depends(get_db),
    current_user: UserResponse = Security(get_current_active_user, scopes=get_required_scopes_for_route("/warehouses"))
):
    """批量导入仓库数据（JSON格式，需要BASE-edit权限）"""
    
    try:
        if not warehouses_data:
            raise HTTPException(status_code=400, detail="请提供要导入的仓库数据")
        
        # 获取仓库导入配置
        config = get_import_config('warehouse')
        if not config:
            raise HTTPException(status_code=400, detail="不支持的实体类型: warehouse")
        
        # 转换数据格式以匹配验证函数期望的格式
        formatted_data = []
        for i, warehouse in enumerate(warehouses_data, start=2):  # 从第2行开始（第1行为标题）
            entity_data = {
                'warehouse_name': warehouse.get('warehouse_name', '') or '',
                'warehouse_city': warehouse.get('warehouse_city', '') or '',
                'warehouse_address': warehouse.get('warehouse_address', '') or '',
                'warehouse_manager': warehouse.get('warehouse_manager', '') or '',
                'warehouse_contact': warehouse.get('warehouse_contact', '') or '',
                'warehouse_level': warehouse.get('warehouse_level')
            }
            formatted_data.append(entity_data)
        
        # 验证数据
        all_errors = await validate_entity_data(formatted_data, config, db)
        
        # 筛选出合格数据
        valid_data = []
        error_data = []
        
        for i, warehouse_data in enumerate(formatted_data):
            row_index = i + 2  # Excel行号
            has_error = any(error.row_index == row_index for error in all_errors)
            if not has_error:
                valid_data.append((row_index, warehouse_data))
            else:
                error_data.append((row_index, warehouse_data))
        
        # 批量插入合格数据
        success_count = 0
        insert_errors = []
        
        # 如果有错误数据，生成错误Excel文件
        error_file_path = None
        error_file_name = None
        if all_errors:
            from utils.error_file_handler import process_import_errors_and_generate_file
            
            # 使用通用函数处理错误文件生成
            # 构造original_rows（JSON导入的情况下需要模拟）
            original_rows = []
            for i, warehouse_data in enumerate(formatted_data):
                row_index = i + 2  # Excel行号
                # 按配置的字段顺序构造行数据
                row_data = tuple(str(warehouse_data.get(field.key, '')) for field in config.template_fields)
                original_rows.append((row_index, row_data))
            
            error_file_path, error_file_name = process_import_errors_and_generate_file(
                all_errors=all_errors,
                original_rows=original_rows,
                config=config,
                entity_key='warehouse',
                username=current_user.username
            )
        
        with batch_import_transaction(db):
            for row_index, entity_data in valid_data:
                try:
                    # 插入前再次检查唯一性（防止并发导入时的重复）
                    existing_warehouse = db.exec(
                        select(Warehouse).where(
                            Warehouse.warehouse_name == entity_data['warehouse_name'].strip(),
                            Warehouse.is_delete != True
                        )
                    ).first()
                    
                    if existing_warehouse:
                        from schemas.common.import_schemas import ImportError
                        insert_errors.append(ImportError(
                            row_index=row_index,
                            field='warehouse_name',
                            error_message=f'仓库:仓库名称:"{entity_data["warehouse_name"]}"在导入过程中与现有数据冲突',
                            raw_data=entity_data
                        ))
                        continue
                    
                    # 创建仓库记录
                    warehouse_data = {
                        'warehouse_name': entity_data.get('warehouse_name', '').strip(),
                        'warehouse_city': entity_data.get('warehouse_city', '').strip() or None,
                        'warehouse_address': entity_data.get('warehouse_address', '').strip() or None,
                        'warehouse_manager': entity_data.get('warehouse_manager', '').strip() or None,
                        'warehouse_contact': entity_data.get('warehouse_contact', '').strip() or None,
                        'warehouse_level': entity_data.get('warehouse_level') or None,
                        'creator': current_user.username
                    }
                    
                    db_warehouse = Warehouse(**warehouse_data)
                    db.add(db_warehouse)
                    success_count += 1
                    
                except Exception as e:
                    from schemas.common.import_schemas import ImportError
                    insert_errors.append(ImportError(
                        row_index=row_index,
                        field='database',
                        error_message=f"仓库:数据库:插入数据库时出错: {str(e)}",
                        raw_data=entity_data
                    ))
        
        # 合并所有错误
        all_errors.extend(insert_errors)
        
        # 构建返回结果
        result = WarehouseBatchImportResult(
            total_count=len(warehouses_data),
            success_count=success_count,
            error_count=len(all_errors),
            errors=all_errors,
            import_time=datetime.now(),
            has_error_file=error_file_path is not None,
            error_file_name=error_file_name
        )
        
        return result
        
    except HTTPException:
        raise
    except Exception as e:
        # 记录详细错误信息
        import traceback
        error_detail = f"批量导入失败: {str(e)}"
        print(f"Import error: {error_detail}")
        print(f"Traceback: {traceback.format_exc()}")
        
        # 返回错误结果
        from schemas.common.import_schemas import ImportError
        return WarehouseBatchImportResult(
            total_count=0,
            success_count=0,
            error_count=1,
            errors=[ImportError(
                row_index=1,
                field="system",
                error_message=error_detail,
                raw_data={}
            )],
            import_time=datetime.now(),
            has_error_file=False
        )

async def process_warehouse_import_data(
    raw_data: List[dict],
    config: Any,
    db: Session,
    current_user: UserResponse,
    background_tasks: BackgroundTasks
) -> WarehouseBatchImportResult:
    """处理仓库导入数据的通用逻辑"""
    
    # 验证和转换数据
    valid_data = []
    errors = []
    
    # 获取现有仓库名称用于唯一性检查
    existing_warehouses = get_existing_values(db, Warehouse, "warehouse_name")
    
    # 处理重复的仓库名称（在同一批数据中）
    seen_names = set()
    
    for index, item in enumerate(raw_data, start=1):
        try:
            # 数据验证
            validation_errors = validate_entity_data(
                entity_data=[item],
                config=config,
                db=db
            )
            
            if validation_errors:
                errors.extend(validation_errors)
                continue
            
            # 构建仓库数据
            warehouse_data = build_entity_data(
                row=tuple(item.get(field.key, '') for field in config.template_fields),
                config=config
            )
            # 添加创建者信息
            warehouse_data['creator'] = current_user.username
            
            valid_data.append(warehouse_data)
            seen_names.add(warehouse_data.get("warehouse_name"))
            
        except Exception as e:
            errors.append(ImportError(
                row_index=index,
                field="general",
                error_message=f"仓库:数据处理:数据处理异常: {str(e)}",
                raw_data=item
            ))
    
    # 如果没有有效数据且没有错误，说明数据为空
    if not valid_data and not errors:
        return WarehouseBatchImportResult(
            total_count=len(raw_data),
            success_count=0,
            error_count=0,
            errors=[],
            import_time=datetime.now(),
            has_error_file=False
        )
    
    # 批量插入数据
    success_count = 0
    
    if valid_data:
        try:
            with batch_import_transaction(db):
                success_count = batch_insert_entities(
                    db=db,
                    entity_class=Warehouse,
                    data_list=valid_data
                )
        except Exception as e:
            # 如果批量插入失败，将所有有效数据标记为错误
            for i, data_item in enumerate(valid_data):
                errors.append(ImportError(
                row_index=i + 1,
                field="general",
                error_message=f"仓库:批量插入:批量插入失败: {str(e)}",
                raw_data=data_item
            ))
    
    # 生成错误文件（如果有错误）
    error_file_path = None
    if errors:
        # 构建原始行数据列表
        original_rows = []
        for index, item in enumerate(raw_data, start=1):
            row_data = tuple(str(item.get(field.key, '')) for field in config.template_fields)
            original_rows.append((index, row_data))
        
        from utils.error_file_handler import process_import_errors_and_generate_file
        error_file_path, error_file_name = process_import_errors_and_generate_file(
            all_errors=errors,
            original_rows=original_rows,
            config=config,
            entity_key='warehouse',
            username=current_user.username
        )
    
    return WarehouseBatchImportResult(
        total_count=len(raw_data),
        success_count=success_count,
        error_count=len(errors),
        errors=errors,
        import_time=datetime.now(),
        has_error_file=error_file_path is not None,
        error_file_name=os.path.basename(error_file_path) if error_file_path else None
    )



# 下载仓库导入错误文件
@warehouse_router.get("/download-error-file")
async def download_warehouse_error_file(
    file_name: str = Query(..., description="错误文件名"),
    current_user: UserResponse = Security(get_current_active_user, scopes=get_required_scopes_for_route("/warehouses/download-error-file"))
):
    """下载仓库导入错误数据文件（需要BASE-edit权限）"""
    # 构建默认仓库错误文件目录
    temp_dir = tempfile.gettempdir()
    warehouse_dir = os.path.join(temp_dir, 'warehouse')
    
    # 直接使用文件名（前端已经进行了URL编码，后端直接使用）
    decoded_file_name = file_name
    
    # 构建完整文件路径
    file_path = os.path.join(warehouse_dir, decoded_file_name)
    
    if not os.path.exists(file_path):
        logger.error(f"错误文件不存在: {file_path}")
        raise HTTPException(status_code=404, detail="错误文件不存在或已过期")
    
    try:
        # 读取文件内容 - 使用二进制模式避免编码问题
        with open(file_path, 'rb') as file:
            file_content = file.read()
        
        # 获取文件大小
        file_size = len(file_content)
        
        logger.info(f"文件读取成功: 文件大小={file_size} bytes")
            
        # 检查文件内容是否为空
        if file_size == 0:
            logger.warning(f"文件内容为空: {file_path}")
            raise HTTPException(status_code=404, detail="错误文件为空或损坏")
        
        # 直接返回响应内容
        return Response(
            content=file_content,
            media_type="application/vnd.ms-excel",
            headers={
                "Content-Disposition": f"attachment; filename=\"{decoded_file_name}\"",
                "Content-Length": str(file_size)
            }
        )
        
    except Exception as e:
        logger.error(f"读取文件失败: {file_path}, 错误: {str(e)}", exc_info=True)
        # 提供更详细的错误信息
        error_detail = f"读取文件失败: {str(e)}"
        if "latin-1" in str(e):
            error_detail += "。请检查文件路径是否包含特殊字符。"
        raise HTTPException(status_code=500, detail=error_detail)


@warehouse_router.get("/import-template")
async def download_warehouse_import_template(
    current_user: UserResponse = Security(get_current_active_user, scopes=get_required_scopes_for_route("/warehouses"))
):
    """下载仓库导入模板文件（需要BASE-edit权限）"""
    return await download_import_template('warehouse')

# 获取所有仓库数据（不分页）
@warehouse_router.get("/all", response_model=List[WarehouseResponse])
async def get_all_warehouses(
    db: Session = Depends(get_db),
    _: UserResponse = Security(get_current_active_user, scopes=get_required_scopes_for_route("/warehouses"))
):
    """获取所有仓库数据，不分页不排序"""
    warehouses = db.exec(
        select(Warehouse).where(Warehouse.is_delete != True)
    ).all()
    return warehouses
    
# 获取仓库列表（分页）
@warehouse_router.get("", response_model=WarehousePaginationResult)
def read_warehouses(
    params: WarehouseQueryParams = Depends(),
    db: Session = Depends(get_db),
    current_user: UserResponse = Security(get_current_active_user, scopes=get_required_scopes_for_route("/warehouses"))
):
    """获取仓库列表（需要BASE-read权限）"""
    
    # 构建查询条件
    query = select(Warehouse).where(Warehouse.is_delete != True)
    
    # 处理多关键词搜索
    if params.search and params.search.strip():
        keywords = params.search.strip().split()
        
        # 每个关键词在四个字段中查找，使用AND逻辑
        for keyword in keywords:
            query = query.where(
                getattr(Warehouse, 'warehouse_name').ilike(f"%{keyword}%") |
                getattr(Warehouse, 'warehouse_city').ilike(f"%{keyword}%") |
                getattr(Warehouse, 'warehouse_address').ilike(f"%{keyword}%") |
                getattr(Warehouse, 'warehouse_manager').ilike(f"%{keyword}%")
            )
    
    # 按仓库名筛选（保持原有功能）
    if params.warehouse_name:
        query = query.where(getattr(Warehouse, 'warehouse_name').ilike(f"%{params.warehouse_name}%"))
    
    # 按城市筛选（保持原有功能）
    if params.warehouse_city:
        query = query.where(getattr(Warehouse, 'warehouse_city').ilike(f"%{params.warehouse_city}%"))
    
    # 按负责人筛选（保持原有功能）
    if params.warehouse_manager:
        query = query.where(getattr(Warehouse, 'warehouse_manager').ilike(f"%{params.warehouse_manager}%"))
    
    # 排序
    if params.sort_field == "id":
        if params.sort_asc:
            query = query.order_by(getattr(Warehouse, 'id'))
        else:
            query = query.order_by(getattr(Warehouse, 'id').desc())
    elif params.sort_field == "warehouse_name":
        if params.sort_asc:
            query = query.order_by(getattr(Warehouse, 'warehouse_name'))
        else:
            query = query.order_by(getattr(Warehouse, 'warehouse_name').desc())
    elif params.sort_field == "create_time":
        if params.sort_asc:
            query = query.order_by(getattr(Warehouse, 'create_time'))
        else:
            query = query.order_by(getattr(Warehouse, 'create_time').desc())
    elif params.sort_field == "update_time":
        if params.sort_asc:
            query = query.order_by(getattr(Warehouse, 'update_time'))
        else:
            query = query.order_by(getattr(Warehouse, 'update_time').desc())
    else:
        # 如果没有有效的排序字段，默认按id升序排序
        query = query.order_by(getattr(Warehouse, 'id'))
    
    # 获取总数
    total_query = select(func.count()).select_from(query.subquery())
    total = db.exec(total_query).one()
    
    # 分页
    query = query.offset((params.page - 1) * params.page_size).limit(params.page_size)
    
    # 执行查询
    warehouses = db.exec(query).all()
    
    # 计算总页数
    total_pages = (total + params.page_size - 1) // params.page_size
    
    return {
        "total": total,
        "page": params.page,
        "page_size": params.page_size,
        "total_pages": total_pages,
        "data": warehouses
    }

# 获取仓库统计信息
@warehouse_router.get("/statistics", response_model=WarehouseStatistics)
def get_warehouse_statistics(
    db: Session = Depends(get_db),
    current_user: UserResponse = Security(get_current_active_user, scopes=get_required_scopes_for_route("/warehouses/statistics"))
):
    """获取仓库统计信息（需要BASE-read权限）"""
    
    # 获取总仓库数
    total_count = db.exec(
        select(func.count()).where(Warehouse.is_delete != True)
    ).one()
    
    # 按城市分组统计
    city_stats_query = select(
        Warehouse.warehouse_city,
        func.count().label("count")
    ).where(Warehouse.is_delete != True).group_by(Warehouse.warehouse_city)
    
    city_stats = db.exec(city_stats_query).all()
    warehouses_by_city = [
        {"city": stat[0], "count": stat[1]}
        for stat in city_stats
    ]
    
    return {
        "total_warehouses": total_count,
        "warehouses_by_city": warehouses_by_city
    }

# 获取单个仓库
@warehouse_router.get("/get/{warehouse_id}", response_model=WarehouseResponse)
def read_warehouse(
    warehouse_id: int,
    db: Session = Depends(get_db),
    current_user: UserResponse = Security(get_current_active_user, scopes=get_required_scopes_for_route("/warehouses/get"))
):
    """获取单个仓库信息（需要BASE-read权限）"""
    
    warehouse = db.get(Warehouse, warehouse_id)
    if not warehouse or warehouse.is_delete:
        raise HTTPException(status_code=404, detail="仓库不存在")
    
    return warehouse

# 创建仓库
@warehouse_router.post("", response_model=WarehouseResponse)
def create_warehouse(
    warehouse: WarehouseCreate,
    db: Session = Depends(get_db),
    current_user: UserResponse = Security(get_current_active_user, scopes=get_required_scopes_for_route("/warehouses/new"))
):
    """创建新仓库（需要BASE-edit权限）"""
    
    # 检查仓库名是否已存在
    existing_warehouse = db.exec(
        select(Warehouse).where(
            Warehouse.warehouse_name == warehouse.warehouse_name,
            Warehouse.is_delete != True
        )
    ).first()
    
    if existing_warehouse:
        raise HTTPException(status_code=400, detail="仓库名已存在")
    
    # 创建新仓库
    db_warehouse = Warehouse(
        warehouse_name=warehouse.warehouse_name,
        warehouse_city=warehouse.warehouse_city,
        warehouse_address=warehouse.warehouse_address,
        warehouse_contact=warehouse.warehouse_contact,
        warehouse_manager=warehouse.warehouse_manager,
        creator=current_user.username  # 使用当前登录用户的用户名作为创建人
    )
    
    db.add(db_warehouse)
    db.commit()
    db.refresh(db_warehouse)
    
    return db_warehouse

# 更新仓库
@warehouse_router.put("/update/{warehouse_id}", response_model=WarehouseResponse)
def update_warehouse(
    warehouse_id: int,
    warehouse_update: WarehouseUpdate,
    db: Session = Depends(get_db),
    current_user: UserResponse = Security(get_current_active_user, scopes=get_required_scopes_for_route("/warehouses/update"))
):
    """更新仓库信息（需要BASE-edit权限）"""
    
    db_warehouse = db.get(Warehouse, warehouse_id)
    if not db_warehouse or db_warehouse.is_delete:
        raise HTTPException(status_code=404, detail="仓库不存在")
    
    # 如果更新仓库名，检查是否与其他仓库重名
    if warehouse_update.warehouse_name and warehouse_update.warehouse_name != db_warehouse.warehouse_name:
        existing_warehouse = db.exec(
            select(Warehouse).where(
                Warehouse.warehouse_name == warehouse_update.warehouse_name,
                Warehouse.is_delete != True,
                Warehouse.id != warehouse_id
            )
        ).first()
        
        if existing_warehouse:
            raise HTTPException(status_code=400, detail="仓库名已存在")
    
    # 更新字段
    update_data = warehouse_update.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        if value is not None:
            setattr(db_warehouse, field, value)
    
    db.add(db_warehouse)
    db.commit()
    db.refresh(db_warehouse)
    
    return db_warehouse

# 删除仓库（软删除）
@warehouse_router.delete("/delete/{warehouse_id}")
def delete_warehouse(
    warehouse_id: int,
    db: Session = Depends(get_db),
    current_user: UserResponse = Security(get_current_active_user, scopes=get_required_scopes_for_route("/warehouses/delete"))
):
    """删除仓库（软删除，需要BASE-edit权限）"""
    
    db_warehouse = db.get(Warehouse, warehouse_id)
    if not db_warehouse or db_warehouse.is_delete:
        raise HTTPException(status_code=404, detail="仓库不存在")
    
    # 软删除
    db_warehouse.is_delete = True
    
    db.add(db_warehouse)
    db.commit()
    
    return {"message": "仓库删除成功"}

# 批量删除仓库
@warehouse_router.post("/batch-delete")
def batch_delete_warehouses(
    batch_delete: BatchWarehouseDelete,
    db: Session = Depends(get_db),
    current_user: UserResponse = Security(get_current_active_user, scopes=get_required_scopes_for_route("/warehouses/delete"))
):
    """批量删除仓库（需要BASE-edit权限）"""
    
    success_count = 0
    failed_warehouses = []
    
    for warehouse_id in batch_delete.warehouse_ids:
        try:
            db_warehouse = db.get(Warehouse, warehouse_id)
            if db_warehouse and not db_warehouse.is_delete:
                db_warehouse.is_delete = True
                db.add(db_warehouse)
                success_count += 1
        except:
            failed_warehouses.append(warehouse_id)
    
    db.commit()
    
    return {
        "success_count": success_count,
        "failed_count": len(failed_warehouses),
        "failed_warehouses": failed_warehouses,
        "message": f"成功删除 {success_count} 个仓库"
    }