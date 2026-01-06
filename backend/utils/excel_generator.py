"""
Excel生成器模块
基于openpyxl的Excel生成工具，支持入库单和出库单的Excel文件生成
"""
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, Font
from typing import List, Dict, Any, Optional
import logging
from pathlib import Path
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP

logger = logging.getLogger(__name__)


def _amount_to_chinese(amount: float) -> str:
    """将金额转换为中文大写"""
    if amount == 0:
        return "零元整"
    
    integer_part = int(amount)
    decimal_part = round((amount - integer_part) * 100)
    
    chinese_digits = ['零', '壹', '贰', '叁', '肆', '伍', '陆', '柒', '捌', '玖']
    chinese_units = ['', '拾', '佰', '仟', '万', '拾', '佰', '仟', '亿', '拾', '佰', '仟', '兆']
    
    result = ''
    
    # 整数部分转换
    if integer_part > 0:
        integer_str = str(integer_part)
        for i, digit in enumerate(integer_str):
            digit_int = int(digit)
            unit_index = len(integer_str) - i - 1
            if unit_index < len(chinese_units):
                result += chinese_digits[digit_int] + chinese_units[unit_index]
            else:
                result += str(digit_int)
        result += '元'
    
    # 小数部分转换
    if decimal_part > 0:
        if decimal_part < 10:
            result += chinese_digits[decimal_part] + '角'
        else:
            jiao = decimal_part // 10
            fen = decimal_part % 10
            result += chinese_digits[jiao] + '角'
            if fen > 0:
                result += chinese_digits[fen] + '分'
    else:
        result += '整'
    
    return result


class InboundOrderExcel:
    """入库单Excel生成器"""
    
    def __init__(self, template_path: Optional[str] = None):
        """
        初始化入库单Excel生成器
        
        Args:
            template_path: 模板文件路径（可选，默认使用backend/template/入库单模板.xlsx）
        """
        if template_path is None:
            # 默认使用backend/template目录下的模板
            backend_dir = Path(__file__).parent.parent
            template_path = backend_dir / "template" / "入库单模板.xlsx"
        
        self.template_path = Path(template_path)
        if not self.template_path.exists():
            # 详细的错误信息，帮助调试
            logger.error(f"Excel模板文件不存在: {self.template_path}")
            logger.error(f"当前工作目录: {Path.cwd()}")
            logger.error(f"__file__路径: {Path(__file__)}")
            logger.error(f"backend目录: {backend_dir}")
            raise FileNotFoundError(f"Excel模板文件不存在: {self.template_path}")
        
        self.order_data = None
        self.items_data = None
        self.total_amount = 0.0
    
    def set_order_data(self, order_data: Dict[str, Any]) -> None:
        """
        设置订单数据
        
        Args:
            order_data: 订单数据字典，包含以下字段：
                - order_number: 入库单号
                - supplier_name: 供应商名称
                - inbound_date: 入库日期（字符串格式 YYYY-MM-DD）
                - creator: 创建人（可选）
        """
        self.order_data = order_data
    
    def set_items_data(self, items_data: List[Dict[str, Any]]) -> None:
        """
        设置物品数据
        
        Args:
            items_data: 物品数据列表，每个元素包含以下字段：
                - material_code: 器材编码
                - material_name: 器材名称
                - material_specification: 器材规格（可选）
                - unit: 单位
                - quantity: 数量
                - unit_price: 单价
        """
        self.items_data = items_data
        # 计算总金额 - 使用Decimal避免浮点数精度问题
        self.total_amount = sum(
            Decimal(str(item.get('quantity', 0))) * Decimal(str(item.get('unit_price', 0)))
            for item in items_data
        )
    
    def generate(self, output_path: str) -> bool:
        """
        生成入库单Excel文件
        
        Args:
            output_path: 输出文件路径
            
        Returns:
            是否生成成功
        """
        try:
            # 加载模板文件
            wb = load_workbook(self.template_path)
            ws = wb.active
            logger.info(f"加载模板文件: {self.template_path}")
            
            # 填充基本信息
            self._fill_header_info(ws)
            
            # 填充明细数据
            self._fill_items_data(ws)
            
            # 添加合计行
            self._add_total_row(ws)
            
            # 保存文件
            wb.save(output_path)
            wb.close()
            
            logger.info(f"入库单Excel文件已生成: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"生成入库单Excel失败: {e}")
            return False
    
    def _fill_header_info(self, ws) -> None:
        """填充表头基本信息"""
        if not self.order_data:
            return
        
        # 定义无边框样式
        no_border = Border(
            left=Side(style=None),
            right=Side(style=None),
            top=Side(style=None),
            bottom=Side(style=None)
        )
        
        # 清除第1、2、3行所有单元格的边框
        for row in range(1, 4):
            for col in range(1, 10):
                cell = ws.cell(row=row, column=col)
                cell.border = no_border
        
        # A3-C3 合并，填入入库单号，左对齐，无边框，自动换行，楷体_GB2312、加粗、12号
        ws.merge_cells('A3:C3')
        ws['A3'].value = f"入库单号：{self.order_data.get('order_number', '')}"
        ws['A3'].font = Font(name='楷体_GB2312', size=12, bold=True)
        ws['A3'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        for col in ['A', 'B', 'C']:
            ws[f'{col}3'].border = no_border
        
        # D3-F3 合并，填入入库日期，左对齐，无边框，自动换行
        ws.merge_cells('D3:F3')
        ws['D3'].value = f"入库日期：{self.order_data.get('inbound_date', '')}"
        ws['D3'].font = Font(name='楷体_GB2312', size=12, bold=True)
        ws['D3'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        for col in ['D', 'E', 'F']:
            ws[f'{col}3'].border = no_border
        
        # H3-I3 合并，填入供应商，左对齐，无边框，自动换行
        ws.merge_cells('H3:I3')
        supplier_name = self.order_data.get('supplier_name', '')
        ws['H3'].value = f"{supplier_name}"
        ws['H3'].font = Font(name='楷体_GB2312', size=12, bold=True)
        ws['H3'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        # 设置最小行高35，根据内容长度动态调整
        # H+I列总宽度16.8磅，字体12磅，约可容纳每行8-10个中文字符或16-20个英文字符
        # 统计中英文字符，英文字符按0.5计算宽度
        char_width = sum(1 if ord(c) > 127 else 0.5 for c in supplier_name)
        # 每行约容纳8个宽度单位（保守估计，考虑字体12磅在16.8宽度下）
        estimated_lines = max(1, int(char_width / 8) + (1 if char_width % 8 > 0 else 0))
        calculated_height = max(35, estimated_lines * 20)  # 每行20磅（含行间距）
        ws.row_dimensions[3].height = calculated_height
        for col in ['H', 'I']:
            ws[f'{col}3'].border = no_border
    
    def _fill_items_data(self, ws) -> None:
        """填充明细数据"""
        if not self.items_data:
            return
        
        # 定义明细数据样式：仿宋_GB2312、12号字、有边框、水平和垂直居中
        detail_font = Font(name='仿宋_GB2312', size=12)
        detail_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        detail_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        start_row = 5
        for index, item in enumerate(self.items_data, 1):
            row = start_row + index - 1
            
            # 计算金额 - 使用Decimal避免浮点数精度问题
            quantity = Decimal(str(item.get('quantity', 0)))
            unit_price = Decimal(str(item.get('unit_price', 0)))
            amount = quantity * unit_price
            
            # 填充每一列数据，并应用样式
            columns_data = [
                (1, index),  # 序号
                (2, item.get('material_code', '')),  # 器材编码
                (3, item.get('material_name', '')),  # 器材名称
                (4, item.get('material_specification', '')),  # 器材规格
                (5, item.get('unit', '')),  # 单位
                (6, item.get('quantity', 0)),  # 数量
                (7, item.get('unit_price', 0)),  # 单价
                (8, amount),  # 金额
                (9, '')  # 备注
            ]
            
            for col, value in columns_data:
                cell = ws.cell(row=row, column=col)
                # 单价和金额使用文本格式，避免科学计数法，支持自动换行不显示###
                if col in [7, 8]:  # 单价和金额列
                    # 保持数据库实际精度，只去掉小数点后多余的0
                    if isinstance(value, (int, float, Decimal)):
                        # 使用Decimal避免浮点数精度问题
                        try:
                            # 将数值转为Decimal，然后转为字符串
                            decimal_value = Decimal(str(value))
                            # 转为字符串并去掉末尾0
                            formatted = str(decimal_value)
                            # 如果是小数，去掉末尾的0
                            if '.' in formatted:
                                formatted = formatted.rstrip('0').rstrip('.')
                            cell.value = formatted
                        except Exception:
                            # 如果转换失败，使用原始值
                            cell.value = str(value)
                    else:
                        cell.value = str(value)
                    cell.number_format = '@'  # 文本格式
                else:
                    cell.value = value
                cell.font = detail_font
                cell.alignment = detail_alignment
                cell.border = detail_border
            
            # 设置行高为自动，以适应换行内容
            ws.row_dimensions[row].height = None
    
    def _add_total_row(self, ws) -> None:
        """添加合计行（两行格式）"""
        if not self.items_data:
            return
        
        start_row = 5
        total_row = start_row + len(self.items_data)
        
        # 定义合计行样式：黑体、12号字、不加粗、有边框、自动换行
        total_font = Font(name='黑体', size=12, bold=False)
        total_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # 第一行：合计 | 人民币(大写)XXX | (小写)XXX
        # A列：合计（水平和垂直居中）
        cell = ws.cell(row=total_row, column=1)
        cell.value = '合计'
        cell.font = total_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = total_border
        
        # B-F列合并：人民币(大写) + 金额中文大写（左对齐，垂直居中）
        amount_chinese = _amount_to_chinese(self.total_amount)
        ws.merge_cells(start_row=total_row, start_column=2, end_row=total_row, end_column=6)
        cell = ws.cell(row=total_row, column=2)
        cell.value = f'人民币(大写){amount_chinese}'
        cell.font = total_font
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell.border = total_border
        for col in range(3, 7):
            ws.cell(row=total_row, column=col).border = total_border
        
        # H-I列合并：(小写) + 金额数字（左对齐，垂直居中）
        formatted_total = f"{self.total_amount:.3f}".rstrip('0').rstrip('.') if '.' in f"{self.total_amount:.3f}" else f"{self.total_amount:.3f}"
        ws.merge_cells(start_row=total_row, start_column=7, end_row=total_row, end_column=9)
        cell = ws.cell(row=total_row, column=7)
        cell.value = f'(小写)：{formatted_total}'
        cell.font = total_font
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell.border = total_border
        ws.cell(row=total_row, column=9).border = total_border
        
        # 设置合计行行高为固定35
        ws.row_dimensions[total_row].height = 35
        
        # 第二行：备注 | 空白内容
        total_row_2 = total_row + 1
        # A列：备注（水平和垂直居中）
        cell = ws.cell(row=total_row_2, column=1)
        cell.value = '备注'
        cell.font = total_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = total_border
        
        # B-I列合并：空白（水平和垂直居中）
        ws.merge_cells(start_row=total_row_2, start_column=2, end_row=total_row_2, end_column=9)
        cell = ws.cell(row=total_row_2, column=2)
        cell.value = ''
        cell.font = total_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = total_border
        for col in range(3, 10):
            ws.cell(row=total_row_2, column=col).border = total_border
        
        # 设置备注行行高为固定35
        ws.row_dimensions[total_row_2].height = 35


class OutboundOrderExcel:
    """出库单Excel生成器"""
    
    def __init__(self, template_path: Optional[str] = None):
        """
        初始化出库单Excel生成器
        
        Args:
            template_path: 模板文件路径（可选，默认使用backend/template/出库单模板.xlsx）
        """
        if template_path is None:
            # 默认使用backend/template目录下的模板
            backend_dir = Path(__file__).parent.parent
            template_path = backend_dir / "template" / "出库单模板.xlsx"
        
        self.template_path = Path(template_path)
        if not self.template_path.exists():
            # 详细的错误信息，帮助调试
            logger.error(f"Excel模板文件不存在: {self.template_path}")
            logger.error(f"当前工作目录: {Path.cwd()}")
            logger.error(f"__file__路径: {Path(__file__)}")
            logger.error(f"backend目录: {backend_dir}")
            raise FileNotFoundError(f"Excel模板文件不存在: {self.template_path}")
        
        self.order_data = None
        self.items_data = None
        self.total_amount = 0.0
    
    def set_order_data(self, order_data: Dict[str, Any]) -> None:
        """
        设置订单数据
        
        Args:
            order_data: 订单数据字典，包含以下字段：
                - order_number: 出库单号
                - customer_name: 客户名称
                - outbound_date: 出库日期（字符串格式 YYYY-MM-DD）
                - creator: 创建人（可选）
        """
        self.order_data = order_data
    
    def set_items_data(self, items_data: List[Dict[str, Any]]) -> None:
        """
        设置物品数据
        
        Args:
            items_data: 物品数据列表，每个元素包含以下字段：
                - material_code: 器材编码
                - material_name: 器材名称
                - material_specification: 器材规格（可选）
                - unit: 单位
                - quantity: 数量
                - unit_price: 单价
        """
        self.items_data = items_data
        # 计算总金额 - 使用Decimal避免浮点数精度问题
        self.total_amount = sum(
            Decimal(str(item.get('quantity', 0))) * Decimal(str(item.get('unit_price', 0)))
            for item in items_data
        )
    
    def generate(self, output_path: str) -> bool:
        """
        生成出库单Excel文件
        
        Args:
            output_path: 输出文件路径
            
        Returns:
            是否生成成功
        """
        try:
            # 加载模板文件
            wb = load_workbook(self.template_path)
            ws = wb.active
            logger.info(f"加载模板文件: {self.template_path}")
            
            # 填充基本信息
            self._fill_header_info(ws)
            
            # 填充明细数据
            self._fill_items_data(ws)
            
            # 添加合计行
            self._add_total_row(ws)
            
            # 保存文件
            wb.save(output_path)
            wb.close()
            
            logger.info(f"出库单Excel文件已生成: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"生成出库单Excel失败: {e}")
            return False
    
    def _fill_header_info(self, ws) -> None:
        """填充表头基本信息"""
        if not self.order_data:
            return
        
        # 定义无边框样式
        no_border = Border(
            left=Side(style=None),
            right=Side(style=None),
            top=Side(style=None),
            bottom=Side(style=None)
        )
        
        # 清除第1、2、3行所有单元格的边框
        for row in range(1, 4):
            for col in range(1, 10):
                cell = ws.cell(row=row, column=col)
                cell.border = no_border
        
        # A3-C3 合并，填入出库单号，左对齐，无边框，自动换行，楷体_GB2312、加粗、12号
        ws.merge_cells('A3:C3')
        ws['A3'].value = f"出库单号：{self.order_data.get('order_number', '')}"
        ws['A3'].font = Font(name='楷体_GB2312', size=12, bold=True)
        ws['A3'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        for col in ['A', 'B', 'C']:
            ws[f'{col}3'].border = no_border
        
        # D3-F3 合并，填入出库日期，左对齐，无边框，自动换行
        ws.merge_cells('D3:F3')
        ws['D3'].value = f"出库日期：{self.order_data.get('outbound_date', '')}"
        ws['D3'].font = Font(name='楷体_GB2312', size=12, bold=True)
        ws['D3'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        for col in ['D', 'E', 'F']:
            ws[f'{col}3'].border = no_border
        
        # H3-I3 合并，填入收货单位，左对齐，无边框，自动换行
        ws.merge_cells('H3:I3')
        customer_name = self.order_data.get('customer_name', '')
        ws['H3'].value = f"{customer_name}"
        ws['H3'].font = Font(name='楷体_GB2312', size=12, bold=True)
        ws['H3'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        # 设置最小行高35，根据内容长度动态调整
        # H+I列总宽度16.8磅，字体12磅，约可容纳每行8-10个中文字符或16-20个英文字符
        # 统计中英文字符，英文字符按0.5计算宽度
        char_width = sum(1 if ord(c) > 127 else 0.5 for c in customer_name)
        # 每行约容纳8个宽度单位（保守估计，考虑字体12磅在16.8宽度下）
        estimated_lines = max(1, int(char_width / 8) + (1 if char_width % 8 > 0 else 0))
        calculated_height = max(35, estimated_lines * 20)  # 每行20磅（含行间距）
        ws.row_dimensions[3].height = calculated_height
        for col in ['H', 'I']:
            ws[f'{col}3'].border = no_border
    
    def _fill_items_data(self, ws) -> None:
        """填充明细数据"""
        if not self.items_data:
            return
        
        # 定义明细数据样式：仿宋_GB2312、12号字、有边框、水平和垂直居中
        detail_font = Font(name='仿宋_GB2312', size=12)
        detail_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        detail_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        start_row = 5
        for index, item in enumerate(self.items_data, 1):
            row = start_row + index - 1
            
            # 计算金额 - 使用Decimal避免浮点数精度问题
            quantity = Decimal(str(item.get('quantity', 0)))
            unit_price = Decimal(str(item.get('unit_price', 0)))
            amount = quantity * unit_price
            
            # 填充每一列数据，并应用样式
            columns_data = [
                (1, index),  # 序号
                (2, item.get('material_code', '')),  # 器材编码
                (3, item.get('material_name', '')),  # 器材名称
                (4, item.get('material_specification', '')),  # 器材规格
                (5, item.get('unit', '')),  # 单位
                (6, item.get('quantity', 0)),  # 数量
                (7, item.get('unit_price', 0)),  # 单价
                (8, amount),  # 金额
                (9, '')  # 备注
            ]
            
            for col, value in columns_data:
                cell = ws.cell(row=row, column=col)
                # 单价和金额使用文本格式，避免科学计数法，支持自动换行不显示###
                if col in [7, 8]:  # 单价和金额列
                    # 保持数据库实际精度，只去掉小数点后多余的0
                    if isinstance(value, (int, float, Decimal)):
                        # 使用Decimal避免浮点数精度问题
                        try:
                            # 将数值转为Decimal，然后转为字符串
                            decimal_value = Decimal(str(value))
                            # 转为字符串并去掉末尾0
                            formatted = str(decimal_value)
                            # 如果是小数，去掉末尾的0
                            if '.' in formatted:
                                formatted = formatted.rstrip('0').rstrip('.')
                            cell.value = formatted
                        except Exception:
                            # 如果转换失败，使用原始值
                            cell.value = str(value)
                    else:
                        cell.value = str(value)
                    cell.number_format = '@'  # 文本格式
                else:
                    cell.value = value
                cell.font = detail_font
                cell.alignment = detail_alignment
                cell.border = detail_border
            
            # 设置行高为自动，以适应换行内容
            ws.row_dimensions[row].height = None
    
    def _add_total_row(self, ws) -> None:
        """添加合计行（两行格式）"""
        if not self.items_data:
            return
        
        start_row = 5
        total_row = start_row + len(self.items_data)
        
        # 定义合计行样式：黑体、12号字、有边框、自动换行
        total_font = Font(name='黑体', size=12, bold=True)
        total_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # 第一行：合计 | 人民币(大写)XXX | (小写)XXX
        # A列：合计（水平和垂直居中）
        cell = ws.cell(row=total_row, column=1)
        cell.value = '合计'
        cell.font = total_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = total_border
        
        # B-G列合并：人民币(大写) + 金额中文大写（左对齐，垂直居中）
        amount_chinese = _amount_to_chinese(self.total_amount)
        ws.merge_cells(start_row=total_row, start_column=2, end_row=total_row, end_column=6)
        cell = ws.cell(row=total_row, column=2)
        cell.value = f'人民币(大写){amount_chinese}'
        cell.font = total_font
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell.border = total_border
        for col in range(3, 7):
            ws.cell(row=total_row, column=col).border = total_border
        
        # H-I列合并：(小写) + 金额数字（左对齐，垂直居中）
        formatted_total = f"{self.total_amount:.3f}".rstrip('0').rstrip('.') if '.' in f"{self.total_amount:.3f}" else f"{self.total_amount:.3f}"
        ws.merge_cells(start_row=total_row, start_column=7, end_row=total_row, end_column=9)
        cell = ws.cell(row=total_row, column=7)
        cell.value = f'(小写)：{formatted_total}'
        cell.font = total_font
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell.border = total_border
        ws.cell(row=total_row, column=9).border = total_border
        # 设置合计行行高为固定35
        ws.row_dimensions[total_row].height = 35

        # 第二行：备注 | 空白内容
        total_row_2 = total_row + 1
        # A列：备注（水平和垂直居中）
        cell = ws.cell(row=total_row_2, column=1)
        cell.value = '备注'
        cell.font = total_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = total_border
        
        # B-I列合并：空白（水平和垂直居中）
        ws.merge_cells(start_row=total_row_2, start_column=2, end_row=total_row_2, end_column=9)
        cell = ws.cell(row=total_row_2, column=2)
        cell.value = ''
        cell.font = total_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = total_border
        for col in range(3, 10):
            ws.cell(row=total_row_2, column=col).border = total_border
        # 设置备注行行高为固定35
        ws.row_dimensions[total_row_2].height = 35

def generate_inbound_order_excel(order_data: Dict[str, Any], items_data: List[Dict[str, Any]], 
                                  output_path: str) -> bool:
    """
    生成入库单Excel文件（便捷函数）
    
    Args:
        order_data: 订单数据字典
        items_data: 物品数据列表
        output_path: 输出文件路径
        
    Returns:
        是否生成成功
    """
    excel = InboundOrderExcel()
    excel.set_order_data(order_data)
    excel.set_items_data(items_data)
    return excel.generate(output_path)


def generate_outbound_order_excel(order_data: Dict[str, Any], items_data: List[Dict[str, Any]], 
                                   output_path: str) -> bool:
    """
    生成出库单Excel文件（便捷函数）
    
    Args:
        order_data: 订单数据字典
        items_data: 物品数据列表
        output_path: 输出文件路径
        
    Returns:
        是否生成成功
    """
    excel = OutboundOrderExcel()
    excel.set_order_data(order_data)
    excel.set_items_data(items_data)
    return excel.generate(output_path)
